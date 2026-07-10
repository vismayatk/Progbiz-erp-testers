"""Generate CRM_Test_Cases.xlsx — 500+ detailed, page-/field-wise test cases."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

COLS = ["Module","Page Name","Test Case ID","Test Scenario","Test Case Description",
        "Preconditions","Test Steps","Test Data","Expected Result","Priority",
        "Test Type","Status","Remarks"]

rows = []
def add(module, page, scen, desc, pre, steps, data, exp, prio, ttype, remarks=""):
    rows.append([module, page, "", scen, desc, pre, steps, data, exp, prio, ttype, "Not Executed", remarks])

def addc(module, page, pre, c):
    # c = (scen, desc, steps, data, exp, prio, ttype)
    add(module, page, c[0], c[1], pre, c[2], c[3], c[4], c[5], c[6])

LOGIN_PRE = "User is on the Login page; valid tenant exists."
ENQ_PRE   = "User is logged in; New Enquiry page is open."

# ============================================================ MODULE 1: LOGIN
M, P = "Login", "Login Page"
login_cases = [
 ("Valid login","Login with valid company/user/password","Enter valid creds → Login","Company=Lesol_dev, User=admin, Pass=123","Redirects to /home; session created","High","Positive"),
 ("Invalid password","Login with wrong password","Valid user, wrong password → Login","User=admin, Pass=wrong","Error 'invalid credentials'; stays on /login","High","Negative"),
 ("Invalid username","Login with non-existent user","Bad username → Login","User=nouser, Pass=123","Error; not logged in","High","Negative"),
 ("Invalid company code","Login with wrong company code","Bad company → Login","Company=bad_co","Error; not logged in","High","Negative"),
 ("Empty username","Submit with username blank","Blank user, valid pass → Login","User='' ","Required-field validation","High","Negative"),
 ("Empty password","Submit with password blank","Valid user, blank pass → Login","Pass='' ","Required-field validation","High","Negative"),
 ("Both fields empty","Submit form empty","Leave all blank → Login","(empty)","Validation on all required fields","Medium","Negative"),
 ("Whitespace only","Username/password spaces only","Enter '   ' → Login","User='   '","Trimmed; treated as empty; validation","Low","Negative"),
 ("Username max length","Boundary at max length","Enter 256-char username","256 chars","Limited to max; no overflow","Medium","Boundary"),
 ("Password min length","Below min length","Enter 1-char password","Pass=a","Min-length validation if enforced","Medium","Boundary"),
 ("Password max length","Above max length","Enter very long password","1000 chars","Handled; no crash","Low","Boundary"),
 ("Special characters","Creds with special chars","Enter !@#$%^&* → Login","User=a!@#","Handled; correct auth result","Medium","Validation"),
 ("Case sensitivity","Username/pass case check","Enter ADMIN vs admin","User=ADMIN","Auth respects configured case rules","Medium","Functional"),
 ("Password masking","Password shows as dots","Type password","Pass=123","Characters masked by default","High","UI"),
 ("Show password icon","Toggle reveal/hide","Click eye icon","-","Password text toggles visible/hidden","Medium","UI"),
 ("Remember Me checked","Persisted session","Check Remember Me → Login → reopen browser","-","Session/login remembered per policy","Medium","Functional"),
 ("Remember Me unchecked","Session not persisted","Uncheck → Login → reopen","-","Requires re-login","Low","Functional"),
 ("Forgot Password link","Navigates to reset","Click Forgot Password","-","Opens reset flow","Medium","Functional"),
 ("Copy-paste credentials","Paste into fields","Paste user/pass → Login","-","Accepted unless paste blocked by policy","Low","Functional"),
 ("Copy from password field","Copy-restriction","Try copy password value","-","Copy blocked/masked per policy","Low","Security"),
 ("SQL injection","Auth bypass attempt","Enter ' OR '1'='1' --","' OR '1'='1","Rejected; no bypass; sanitized","High","Security"),
 ("XSS in fields","Script payload","Enter <script>alert(1)</script>","<script>alert(1)</script>","Escaped; no script execution","High","Security"),
 ("Brute force","Repeated wrong attempts","Submit wrong pass 5-10 times","-","Throttle/captcha/lockout triggered","High","Security"),
 ("Locked account","Login to locked user","Use locked account","-","'Account locked' message; blocked","High","Negative"),
 ("Inactive user","Login to inactive user","Use inactive account","-","Blocked with proper message","High","Negative"),
 ("Deleted user","Login to deleted user","Use deleted account","-","Blocked; not authenticated","High","Negative"),
 ("Session timeout","Idle session expiry","Login, idle past timeout, act","-","Redirect to /login; session invalid","High","Functional"),
 ("Browser refresh on login","Refresh mid-login","Enter creds, refresh","-","No partial state; form resets safely","Low","Functional"),
 ("Back button after logout","Back navigation","Logout → press Back","-","Cannot access app; redirect to /login","High","Security"),
 ("Multiple tabs/browsers","Same user multiple sessions","Login same user in 2 browsers","-","Per policy: both allowed or 2nd invalidates 1st","Medium","Functional"),
 ("Concurrent users","Many users at once","N users login simultaneously","-","All authenticate; no data mix","Medium","Performance"),
 ("Network interruption","Drop connection during login","Disconnect while submitting","-","Graceful error; retry possible","Medium","Negative"),
 ("Slow internet","Throttled login","Login on slow 3G","-","Loading indicator; no double-submit","Low","Performance"),
 ("Double-click Login","Prevent double submit","Click Login twice fast","-","Single auth request; no duplicate session","Medium","Functional"),
 ("Mobile responsive","Login on mobile","Open on 375px width","-","Layout adapts; fields usable","Medium","UI"),
 ("Tablet responsive","Login on tablet","Open on 768px","-","Layout adapts","Low","UI"),
 ("Cross-browser","Chrome/Firefox/Edge/Safari","Login on each browser","-","Consistent behavior across browsers","Medium","Compatibility"),
 ("Keyboard navigation","Tab order / Enter submits","Tab through fields, press Enter","-","Logical tab order; Enter submits","Low","Accessibility"),
 ("Screen reader labels","Field labels announced","Use screen reader","-","Inputs have accessible labels","Low","Accessibility"),
 ("HTTPS / secure transport","Credentials over TLS","Inspect network","-","Login posted over HTTPS only","High","Security"),
]
for c in login_cases: addc(M,P,LOGIN_PRE,c)

# ============================================================ MODULE 2: CRM LIST
M, P = "CRM List", "Leads / Enquiry List Page"
LIST_PRE = "User logged in; list page has records."
list_cases = [
 ("Grid data load","Records load on open","Open list page","-","Rows render with correct columns","High","Functional"),
 ("Record count","Total count accurate","Read count vs DB","-","Displayed count matches actual","Medium","Data Integrity"),
 ("Pagination next/prev","Navigate pages","Click next/prev","-","Correct page set loads","High","Functional"),
 ("Page size change","Change rows per page","Set 10/25/50/100","-","List re-renders with chosen size","Medium","Functional"),
 ("Default sorting","Initial sort order","Open list","-","Sorted by default column (e.g. date desc)","Low","UI"),
 ("Single column sort","Sort asc/desc","Click a header","-","Rows sort correctly","Medium","Functional"),
 ("Multi-column sort","Secondary sort","Shift-click headers","-","Stable multi-key sort","Low","Functional"),
 ("Sticky headers","Header stays on scroll","Scroll list","-","Header remains visible","Low","UI"),
 ("Exact search","Full value search","Search exact name","Existing name","Only exact matches","High","Positive"),
 ("Partial search","Substring search","Search part of name","'Test'","All containing rows","High","Positive"),
 ("Case-insensitive search","Lower/upper","Search 'test' vs 'TEST'","-","Same results","Medium","Functional"),
 ("Special char search","Symbols in search","Search '%','_','''","-","No crash; escaped","Medium","Security"),
 ("Empty search","Blank query","Clear search","-","Full list restored","Low","Functional"),
 ("Search with spaces","Leading/trailing spaces","Search ' test '","-","Trimmed; matches","Low","Validation"),
 ("Search then filter","Combine","Search + apply filter","-","Both criteria applied","Medium","Functional"),
 ("No-result search","Non-existent term","Search 'zzzzz'","-","'No data' empty state; no error","Medium","Negative"),
 ("Date filter","Filter by date range","Set from/to","-","Rows within range only","High","Functional"),
 ("Future date filter","Range in future","From/to in future","-","Empty/valid handling; no error","Low","Boundary"),
 ("Invalid date range","From > To","From later than To","-","Blocked/validation message","Medium","Negative"),
 ("Lead source filter","Filter by source","Select a source","-","Matching rows only","Medium","Functional"),
 ("Lead status filter","Filter by status","Select status/tab","-","Matching rows only","Medium","Functional"),
 ("Assigned user filter","Filter by executive","Select user","-","Matching rows only","Medium","Functional"),
 ("Branch filter","Filter by branch","Select branch","-","Matching rows only","Medium","Functional"),
 ("Multiple filters","Combine filters","Apply several filters","-","AND of all filters","High","Functional"),
 ("Clear filter","Clear one filter","Click clear","-","That filter removed; list updates","Medium","Functional"),
 ("Reset filters","Reset all","Click Reset","-","All filters cleared; full list","Medium","Functional"),
 ("Excel export","Export to xlsx","Click Export Data","-","Valid .xlsx with correct columns/rows","High","Functional"),
 ("Export filtered","Export current filter","Filter then export","-","Only filtered rows exported","High","Data Integrity"),
 ("Export all","Export full dataset","Export without filter","-","All rows exported","Medium","Functional"),
 ("Export selected","Export checked rows","Select rows → export","-","Only selected rows","Medium","Functional"),
 ("Large export","Export many rows","Export 10k+","-","Completes; no timeout/crash","Medium","Performance"),
 ("Empty export","Export empty list","Filter to 0 → export","-","Header-only file / message","Low","Negative"),
 ("Export column validation","Columns match grid","Open exported file","-","Columns and order correct","Medium","Data Integrity"),
 ("View action","Open record","Click row/view","-","Detail/overview opens","High","Functional"),
 ("Edit action","Edit record","Click edit","-","Edit form prefilled","High","Functional"),
 ("Delete action","Delete record","Click delete → confirm","-","Row removed after confirm","High","Functional"),
 ("Delete cancel","Cancel deletion","Delete → cancel","-","Row retained","Medium","Negative"),
 ("Bulk select","Select multiple","Check several / select-all","-","Selection count updates","Medium","Functional"),
 ("Bulk delete","Delete many","Select rows → bulk delete","-","All selected removed (with confirm)","High","Functional"),
 ("Bulk assign","Assign many","Select rows → assign user","-","Assignee updated for all","High","Functional"),
 ("Bulk update","Update many","Select → bulk edit field","-","Field updated for all","Medium","Functional"),
 ("Refresh persists filter","Refresh page","Apply filter → refresh","-","Per design: filters persist or reset cleanly","Low","Functional"),
]
for c in list_cases: addc(M,P,LIST_PRE,c)

# ============================================================ MODULE 3: NEW ENQUIRY (field-wise)
M, P = "New Enquiry", "New Enquiry / Add Lead Page"
# field, category, mandatory, maxlen
FIELDS = [
 ("Enquiry Number","auto",False,None),
 ("Customer Name","text",True,100),
 ("Company Name","text",False,150),
 ("Mobile Number","phone",True,10),
 ("Alternate Number","phone",False,10),
 ("Email","email",False,254),
 ("Country","dropdown",False,None),
 ("State","dropdown_dep",False,None),
 ("City","text",False,100),
 ("Address","text",False,250),
 ("Pincode","number",False,6),
 ("Lead Source","dropdown",False,None),
 ("Lead Status","dropdown",False,None),
 ("Branch","dropdown",True,None),
 ("Assigned User","dropdown",True,None),
 ("Priority","dropdown",False,None),
 ("Product","dropdown",False,None),
 ("Service","dropdown",False,None),
 ("Description","text",False,500),
 ("Remarks","text",False,500),
 ("Follow-up Date","date",False,None),
 ("Follow-up Time","time",False,None),
 ("Attachment Upload","file",False,None),
 ("Tags","text",False,100),
 ("Custom Fields","text",False,100),
]

CUR_OPEN = "Open New Enquiry"
CUR_PRE  = ENQ_PRE
def sc(f, name, desc, data, exp, prio, ttype, action=None):
    act = action or f"Enter test data into '{f}'"
    steps = f"1. {CUR_OPEN}. 2. Locate '{f}'. 3. {act}. 4. Click Save."
    return (name, desc, CUR_PRE, steps, data, exp, prio, ttype)

def field_cases(f, cat, mand, maxlen):
    out=[]
    if cat=="auto":
        out += [
         sc(f,f"{f} - auto generated","Number auto-populates on page open","-","Unique sequential number shown (e.g. ENQ-###)","High","Positive"),
         sc(f,f"{f} - read only","User cannot edit the number","Type into field","Field is non-editable","Medium","Negative"),
         sc(f,f"{f} - uniqueness","No two enquiries share a number","Create two enquiries","Each gets a distinct number","High","Data Integrity"),
         sc(f,f"{f} - sequence after save","Number increments","Save then add new","Next number follows sequence","Medium","Functional"),
        ]
        return out
    if cat=="text":
        if mand: out.append(sc(f,f"{f} - empty mandatory","Required validation",". (leave blank)","'<field> is required'; not saved","High","Negative"))
        out += [
         sc(f,f"{f} - valid value","Accepts valid input","Valid text","Accepted; saved","High","Positive"),
         sc(f,f"{f} - max length","Boundary at max ({maxlen})",f"{maxlen} chars","Accepted up to limit","Medium","Boundary"),
         sc(f,f"{f} - exceed max","Above max length",f"{(maxlen or 100)+1} chars","Blocked/truncated at {maxlen}","Medium","Boundary"),
         sc(f,f"{f} - leading spaces","Trim leading","'   value'","Trimmed before save","Low","Validation"),
         sc(f,f"{f} - trailing spaces","Trim trailing","'value   '","Trimmed before save","Low","Validation"),
         sc(f,f"{f} - special chars","Symbols handled","!@#$%^&*()","Handled; no break","Medium","Validation"),
         sc(f,f"{f} - emoji","Emoji input","😀🚀","Handled/rejected gracefully","Low","Negative"),
         sc(f,f"{f} - numbers only","Numeric in text field","1234567890","Accepted/validated per rule","Low","Validation"),
         sc(f,f"{f} - copy paste","Paste value","paste 'Test'","Accepted","Low","Functional"),
         sc(f,f"{f} - SQL injection","Injection payload","' OR 1=1 --","Sanitized; stored safely","High","Security"),
         sc(f,f"{f} - XSS injection","Script payload","<script>alert(1)</script>","Escaped; no execution on view","High","Security"),
        ]
        if f in ("Customer Name","Company Name","Tags"):
            out.append(sc(f,f"{f} - duplicate","Duplicate value handling","existing value","Allowed or flagged per business rule","Medium","Negative"))
        return out
    if cat=="email":
        out += [
         sc(f,f"{f} - valid email","Accepts valid email","user@example.com","Accepted","High","Positive"),
         sc(f,f"{f} - missing @","Invalid format","userexample.com","Validation error","High","Negative"),
         sc(f,f"{f} - missing domain","Invalid format","user@","Validation error","Medium","Negative"),
         sc(f,f"{f} - spaces","Spaces in email","'a b@x.com'","Rejected/trimmed","Low","Validation"),
         sc(f,f"{f} - uppercase","Case handling","USER@EXAMPLE.COM","Accepted; normalized","Low","Validation"),
         sc(f,f"{f} - max length","Long email","254 chars","Accepted up to limit","Low","Boundary"),
         sc(f,f"{f} - duplicate email","Unique enforcement","existing email","Rejected if unique enforced","High","Negative"),
         sc(f,f"{f} - SQL/XSS","Injection in email","a@x.com'<script>","Sanitized","High","Security"),
         sc(f,f"{f} - empty optional","Blank allowed",". (blank)","Saved (field optional)","Low","Positive"),
        ]
        return out
    if cat=="phone":
        if mand: out.append(sc(f,f"{f} - empty mandatory","Required validation",". (blank)","Validation; not saved","High","Negative"))
        out += [
         sc(f,f"{f} - valid 10 digit","Accepts valid number","9876543210","Accepted","High","Positive"),
         sc(f,f"{f} - less than 10","Below length","98765","Length validation","High","Boundary"),
         sc(f,f"{f} - more than 10","Above length","987654321012","Blocked/validation","Medium","Boundary"),
         sc(f,f"{f} - alphabets","Non-numeric","98abc43210","Rejected","High","Negative"),
         sc(f,f"{f} - special chars","Symbols","+91-98765","Per format rule (+/-) handled","Medium","Validation"),
         sc(f,f"{f} - leading zeros","Zeros","0000000000","Per rule","Low","Boundary"),
         sc(f,f"{f} - duplicate phone","Existing number","existing phone","'phone already exists' if unique","High","Negative"),
         sc(f,f"{f} - SQL/XSS","Injection","' OR 1=1","Sanitized","High","Security"),
        ]
        return out
    if cat=="number":
        out += [
         sc(f,f"{f} - valid","Valid numeric","560001","Accepted","High","Positive"),
         sc(f,f"{f} - non numeric","Letters","abc123","Rejected","High","Negative"),
         sc(f,f"{f} - below length","Short","123","Length validation","Medium","Boundary"),
         sc(f,f"{f} - above length","Long","1234567","Blocked","Medium","Boundary"),
         sc(f,f"{f} - special chars","Symbols","56-001","Rejected/handled","Low","Validation"),
         sc(f,f"{f} - empty optional","Blank",". (blank)","Saved (optional)","Low","Positive"),
        ]
        return out
    if cat in ("dropdown","dropdown_dep"):
        out += [
         sc(f,f"{f} - select valid","Pick an option","Select option","Value set; saved","High","Positive",f"Select a value in '{f}'"),
         sc(f,f"{f} - options load","List populates","Open dropdown","All configured options shown","Medium","Functional",f"Open '{f}'"),
         sc(f,f"{f} - searchable","Type-ahead","Type to filter","Matching options shown","Low","UI",f"Type in '{f}'"),
         sc(f,f"{f} - default value","Default selection","Open page","Correct default (e.g. 'Choose')","Low","UI",f"Inspect '{f}'"),
        ]
        if mand: out.append(sc(f,f"{f} - no selection mandatory","Required validation","Leave at 'Choose'","Validation; not saved","High","Negative",f"Do not select '{f}'"))
        if cat=="dropdown_dep":
            out.append(sc(f,f"{f} - dependency","Depends on parent (Country)","Change Country","'{f}' options refresh accordingly","Medium","Functional",f"Change parent then open '{f}'"))
        return out
    if cat=="date":
        out += [
         sc(f,f"{f} - valid future","Future date accepted","Tomorrow","Accepted","High","Positive"),
         sc(f,f"{f} - current date","Today","Today","Accepted per rule","Medium","Positive"),
         sc(f,f"{f} - past date","Past date","Yesterday","Rejected: 'must be future'","High","Negative"),
         sc(f,f"{f} - invalid format","Bad format","32/13/2026","Rejected","Medium","Negative"),
         sc(f,f"{f} - weekend","Weekend date","Saturday","Allowed/flagged per policy","Low","Validation"),
         sc(f,f"{f} - holiday","Holiday date","Public holiday","Allowed/flagged per policy","Low","Validation"),
         sc(f,f"{f} - empty optional","Blank / Not Required",". (blank or tick Not Required)","Saved","Low","Positive"),
        ]
        return out
    if cat=="time":
        out += [
         sc(f,f"{f} - valid time","Valid time","14:30","Accepted","High","Positive"),
         sc(f,f"{f} - invalid time","Bad time","25:99","Rejected","Medium","Negative"),
         sc(f,f"{f} - past time today","Past time same day","Earlier today","Per rule rejected/flagged","Medium","Negative"),
         sc(f,f"{f} - empty","Blank",". (blank)","Per rule","Low","Validation"),
        ]
        return out
    if cat=="file":
        out += [
         sc(f,f"{f} - valid file","Allowed type/size","upload PDF 1MB","Uploaded; shown","High","Positive"),
         sc(f,f"{f} - unsupported ext","Blocked extension","upload .exe","Rejected with message","High","Negative"),
         sc(f,f"{f} - oversized","Above size limit","upload 50MB","Rejected: size limit","High","Boundary"),
         sc(f,f"{f} - multiple files","Multi-upload","upload 3 files","All listed (if allowed)","Medium","Functional"),
         sc(f,f"{f} - corrupted file","Bad file","upload corrupt PDF","Handled gracefully","Medium","Negative"),
        ]
        return out
    return out

for f,cat,mand,maxlen in FIELDS:
    for case in field_cases(f,cat,mand,maxlen):
        add(M,P,case[0],case[1],case[2],case[3],case[4],case[5],case[6],case[7])

# ============================================================ MODULE 4: SAVE ENQUIRY
M,P="Save Enquiry","New Enquiry Page"
save_cases=[
 ("Save complete data","All fields valid","Fill all valid → Save","valid set","Saved; success; redirect to overview","High","Positive"),
 ("Save mandatory only","Only required fields","Fill required → Save","required set","Saved successfully","High","Positive"),
 ("Save partial data","Some optional missing","Fill some → Save","partial","Saved if required present","Medium","Positive"),
 ("Save without data","Empty form","Click Save empty","-","Validation on all required","High","Negative"),
 ("Save duplicate enquiry","Duplicate customer/phone","Save with existing phone","existing phone","Duplicate handling/rejection","High","Negative"),
 ("Save without item","No line item","Save without adding item","-","'Please choose at least one item'","High","Negative"),
 ("Double-click Save","Prevent double submit","Click Save twice","-","Single record; no duplicate","High","Functional"),
 ("Multiple rapid saves","Spam Save","Click Save repeatedly","-","Only one enquiry created","Medium","Negative"),
 ("Network disconnect on save","Offline mid-save","Disconnect then Save","-","Error shown; no partial save","Medium","Negative"),
 ("Session expired on save","Expired session","Idle then Save","-","Redirect to login; data preserved/warned","High","Negative"),
 ("Server error on save","5xx response","Trigger server error","-","Friendly error; not a blank screen","Medium","Negative"),
 ("DB failure on save","Backend exception","Force backend error","-","Clear error (e.g. 'Oops Error Code')","High","Negative"),
 ("Browser refresh during save","Refresh mid-save","Refresh while saving","-","No duplicate/partial record","Medium","Negative"),
 ("Auto enquiry number on save","Number assigned","Save valid enquiry","-","Unique enquiry number generated","High","Data Integrity"),
 ("Attachment saved","File persists","Save with attachment","PDF","Attachment stored & retrievable","Medium","Functional"),
 ("Cancel after fill","Discard","Fill → Cancel","-","No record saved; form closes","Medium","Functional"),
]
for c in save_cases: addc(M,P,ENQ_PRE,c)

# ============================================================ MODULE 5: FOLLOW-UP CREATION
M,P="Follow-up","Enquiry Overview / Follow-up Modal"
FU_PRE="An enquiry/lead exists and is open."
fu_cases=[
 ("Create follow-up","Valid follow-up","Followup → status + quality → Save","status=Interested","Saved; status updated; row in history","High","Positive"),
 ("Follow-up without enquiry","No parent lead","Try create without lead context","-","Blocked; requires enquiry","High","Negative"),
 ("Valid future date","Future next follow-up","Set tomorrow","tomorrow","Accepted","High","Positive"),
 ("Invalid date","Bad date","Enter invalid","99/99","Rejected","Medium","Negative"),
 ("Past date","Past next follow-up","Yesterday","yesterday","'must be future' error","High","Negative"),
 ("Current date","Today","Today","today","Accepted per rule","Medium","Positive"),
 ("Future date","Far future","+30 days","+30d","Accepted","Low","Positive"),
 ("Weekend date","Weekend","Saturday","Sat","Allowed/flagged","Low","Validation"),
 ("Holiday date","Holiday","Public holiday","holiday","Allowed/flagged","Low","Validation"),
 ("Valid time","Valid time","14:00","14:00","Accepted","Medium","Positive"),
 ("Invalid time","Bad time","26:00","26:00","Rejected","Medium","Negative"),
 ("No status selected","Missing status","Save without status","-","Validation on status","High","Negative"),
 ("No lead quality","Missing required quality","Status set, quality blank","-","Validation on Lead Quality","Medium","Negative"),
 ("Multiple follow-ups","Several on one lead","Add 3 follow-ups","-","All recorded in history","Medium","Functional"),
 ("Follow-up after converted","Lead converted","Convert → try follow-up","-","Followup unavailable by design","Medium","Negative"),
 ("Follow-up after closed","Closed lead","Close → try follow-up","-","Blocked/limited per rule","Medium","Negative"),
 ("Not Required toggle","Skip next date","Tick 'Not Required'","-","Saves without next date","Low","Positive"),
]
for c in fu_cases: addc(M,P,FU_PRE,c)

# ============================================================ MODULE 6: FOLLOW-UP DATE CHANGE
M,P="Follow-up Reschedule","Follow-up / Reschedule"
fd_cases=[
 ("Reschedule follow-up","Move date","Edit next follow-up date","new date","Updated; reflected in lists","High","Positive"),
 ("Move to future","Future reschedule","Set future date","+5d","Accepted","High","Positive"),
 ("Move to current","Today","Set today","today","Accepted per rule","Medium","Positive"),
 ("Move to past","Past reschedule","Set past date","-1d","Rejected","High","Negative"),
 ("Holiday validation","Holiday target","Set holiday","holiday","Allowed/flagged","Low","Validation"),
 ("Weekend validation","Weekend target","Set weekend","Sat","Allowed/flagged","Low","Validation"),
 ("Time conflict","Overlapping slot","Set conflicting time","-","Warned per policy","Low","Validation"),
 ("User availability","Assignee busy","Reschedule to busy slot","-","Warned/allowed per policy","Low","Functional"),
 ("Bulk reschedule","Many leads","Select leads → change date","-","All updated","Medium","Functional"),
 ("Audit log","Change tracked","Reschedule then view log","-","Old/new date logged with user/time","Medium","Data Integrity"),
 ("Notification","User notified","Reschedule","-","Notification raised","Low","Functional"),
 ("Email notification","Email sent","Reschedule","-","Email triggered (if configured)","Low","Functional"),
 ("SMS notification","SMS sent","Reschedule","-","SMS triggered (if configured)","Low","Functional"),
 ("WhatsApp notification","WA sent","Reschedule","-","WA message triggered (if configured)","Low","Functional"),
 ("Calendar sync","Calendar updated","Reschedule","-","Linked calendar event updates","Low","Functional"),
]
for c in fd_cases: addc(M,P,"A follow-up exists for a lead.",c)

# ============================================================ MODULE 7: ATTACHMENT
M,P="Attachment","Enquiry Attachment"
at_cases=[
 ("Generic file upload","Allowed file","Upload valid file","file.txt","Uploaded","High","Positive"),
 ("Image upload","JPG/PNG","Upload image","img.png","Uploaded; preview if supported","Medium","Positive"),
 ("PDF upload","PDF","Upload pdf","doc.pdf","Uploaded","High","Positive"),
 ("Excel upload","XLSX","Upload excel","sheet.xlsx","Uploaded","Medium","Positive"),
 ("Word upload","DOCX","Upload word","doc.docx","Uploaded","Low","Positive"),
 ("Large file","Above size limit","Upload 50MB","big.pdf","Rejected: size limit","High","Boundary"),
 ("Corrupted file","Broken file","Upload corrupt","corrupt.pdf","Handled; error message","Medium","Negative"),
 ("Virus/EICAR file","Malicious file","Upload EICAR test","eicar.com","Blocked by scan (if configured)","High","Security"),
 ("Unsupported extension","Blocked type","Upload .exe","app.exe","Rejected","High","Negative"),
 ("Multiple files","Batch upload","Upload several","3 files","All accepted (if allowed)","Medium","Functional"),
 ("Duplicate file","Same file twice","Upload same name","dup.pdf","Handled/renamed/flagged","Low","Negative"),
 ("Zero-byte file","Empty file","Upload 0KB","empty.pdf","Rejected/handled","Low","Boundary"),
 ("Filename special chars","Odd name","Upload 'a b#@.pdf'","-","Handled safely","Low","Validation"),
 ("Download attachment","Retrieve","Open saved attachment","-","Correct file downloads","Medium","Functional"),
 ("Delete attachment","Remove","Delete uploaded file","-","Removed from record","Medium","Functional"),
]
for c in at_cases: addc(M,P,"An enquiry is open.",c)

# ============================================================ MODULE 8: SECURITY
M,P="Security","Application-wide"
sec_cases=[
 ("SQL injection","DB attack via inputs","Inject SQL in fields/params","' OR 1=1 --","Sanitized; no DB leak","High","Security"),
 ("Stored/Reflected XSS","Script injection","Inject <script> in fields","<script>alert(1)</script>","Escaped on render","High","Security"),
 ("CSRF","Cross-site request","Forge state-changing request","-","Token required; forged request blocked","High","Security"),
 ("Privilege escalation","Low→high access","Access admin features as normal user","-","Denied (403)","High","Security"),
 ("Direct URL access","IDOR/forced browse","Open /enquiry-overview/{otherId}","-","Authorization enforced; denied","High","Security"),
 ("Session hijacking","Stolen token reuse","Reuse session cookie","-","Bound/invalidated; cannot reuse","High","Security"),
 ("Unauthorized access","No login","Open app URL logged out","-","Redirect to /login","High","Security"),
 ("Parameter tampering","Modify IDs/amounts","Edit request params","-","Server validates; rejects tampering","High","Security"),
 ("File upload security","Malicious upload","Upload script/exe","-","Blocked; not executed","High","Security"),
 ("API security / authz","API without/with wrong token","Call API endpoints","-","401/403 enforced; rate-limited","High","Security"),
 ("Sensitive data exposure","PII in responses/logs","Inspect responses","-","No secrets/PII leaked; HTTPS","High","Security"),
 ("Password storage","Hashing","Inspect stored password","-","Hashed+salted; never plaintext","High","Security"),
]
for c in sec_cases: addc(M,P,"Application is accessible.",c)

# ============================================================ MODULE 9: PERFORMANCE
M,P="Performance","List / Search / Export / Save"
perf_cases=[
 ("Load 100 records","Small dataset","Open list with 100","-","Loads < 2s","Medium","Performance"),
 ("Load 1000 records","Medium dataset","Open list with 1k","-","Loads within acceptable time","Medium","Performance"),
 ("Load 10000 records","Large dataset","Open list with 10k (paged)","-","Pagination keeps it responsive","High","Performance"),
 ("Load 50000 records","Very large","Open with 50k","-","No crash; server-side paging","High","Performance"),
 ("Search performance","Search on large data","Search in 50k","-","Results < 3s","High","Performance"),
 ("Filter performance","Filter large data","Apply filters on 50k","-","Acceptable response","Medium","Performance"),
 ("Export performance","Export large data","Export 10k+","-","Completes without timeout","Medium","Performance"),
 ("Save performance","Save under load","Save during heavy use","-","Save < 3s","Medium","Performance"),
 ("Concurrent users","Many users","100 concurrent users","-","Stable; acceptable latency","High","Performance"),
 ("Dashboard load","Heavy widgets","Open CRM dashboard","-","All widgets load reasonably","Medium","Performance"),
 ("Memory/leak on navigation","Repeated navigation","Navigate pages 50x","-","No memory growth/slowdown","Low","Performance"),
]
for c in perf_cases: addc(M,P,"Test dataset is prepared.",c)

# ============================================================ MODULE 10: RESPONSIVE / COMPAT
M,P="Responsive & Compatibility","All pages"
devices=["Desktop (1920)","Laptop (1366)","Tablet (768)","Mobile (375)"]
browsers=["Chrome","Firefox","Edge","Safari"]
for d in devices:
    add(M,P,f"Layout on {d}",f"Render & usability on {d}","Logged in","Open key pages on {d}".format(d=d),d,"Layout adapts; no overflow; controls usable","Medium","UI")
for b in browsers:
    add(M,P,f"Cross-browser on {b}",f"Functional parity on {b}","Logged in",f"Run smoke flow on {b}","-","Consistent behavior; no console errors","Medium","Compatibility")
resp_extra=[
 ("Orientation change","Portrait/landscape","Rotate device","-","Layout reflows correctly","Low","UI"),
 ("Touch interactions","Tap/scroll/swipe","Use on touch device","-","All actions work by touch","Low","UI"),
 ("Zoom 200%","Accessibility zoom","Zoom browser to 200%","-","Content readable; no clipping","Low","Accessibility"),
 ("Slow 3G","Throttled network","Load on slow 3G","-","Loaders shown; usable","Low","Performance"),
 ("Offline mode","No network","Disconnect mid-use","-","Graceful messaging","Low","Negative"),
]
for c in resp_extra: addc(M,P,"User is logged in.",c)

# ============================================================ MORE PAGES (field-wise)
def run_fields(module, page, open_step, pre, fieldlist):
    global CUR_OPEN, CUR_PRE
    CUR_OPEN, CUR_PRE = open_step, pre
    for f,cat,mand,maxlen in fieldlist:
        for case in field_cases(f,cat,mand,maxlen):
            add(module,page,case[0],case[1],case[2],case[3],case[4],case[5],case[6],case[7])

# New Customer modal (opened from the enquiry phone search)
run_fields("New Enquiry","New Customer Modal","Open New Enquiry → phone '+' → New Customer modal",
 "New Enquiry open; New Customer modal shown",[
 ("Customer Level","dropdown",True,None),("Customer Name","text",True,100),
 ("Customer Phone","phone",True,10),("Secondary Phone","phone",False,10),
 ("Email Address","email",False,254),("Place","text",False,100),("Currency","dropdown",False,None),
])
# Create Quotation page
run_fields("Quotation","Create Quotation Page","Open Create Quotation","User logged in; Create Quotation open",[
 ("Customer Name","text",True,100),("Company Name","text",True,150),("Address","text",True,250),
 ("Country","dropdown",False,None),("State","dropdown_dep",False,None),("Contact Person","text",True,100),
 ("Contact No","phone",True,10),("Contact Email","email",True,254),("Tax Reg Status","dropdown",False,None),
 ("Address Type","dropdown",True,None),
])
# Item / Product form
run_fields("Inventory","Item / Product Form","Open Items → New Item","User logged in; Product form open",[
 ("Item Type","dropdown",True,None),("Item Name","text",True,100),("Group","dropdown",False,None),
 ("Category","dropdown",False,None),("UOM Unit","dropdown",False,None),("Brand","dropdown",False,None),
 ("SKU","text",False,50),("Barcode","text",False,50),("Price","number",False,12),("Description","text",False,500),
])

# Lead Transfer (curated)
M,P="Lead Transfer","Bulk Lead Transfer Page"; LT_PRE="User logged in; leads exist."
lt_cases=[
 ("Apply Filters loads leads","Leads list after filter","Apply Filters","-","Matching leads listed with checkboxes","High","Functional"),
 ("Select single lead","Pick one lead","Tick a checkbox","-","Selection count = 1; transfer bar shows","High","Functional"),
 ("Select all leads","Bulk select","Click Select All","-","All visible selected","Medium","Functional"),
 ("Choose executive","Pick target","Open 'Transfer To'","-","Executive list shown","High","Functional"),
 ("Transfer to executive","Reassign","Select lead+exec → Transfer Selected → confirm","-","'Successfully transferred'; assignee updated","High","Positive"),
 ("Verify new assignee","Reassignment check","Re-search lead","-","Current Assignee = target executive","High","Data Integrity"),
 ("Transfer with no lead","Missing selection","Choose exec, transfer 0 selected","-","Disabled / 'select at least one'","High","Negative"),
 ("Transfer with no executive","Missing target","Select lead, no exec → Transfer","-","'Select Executive' validation","High","Negative"),
 ("Empty filter result","No match","Filter to 0 leads","-","Empty state; no error","Medium","Negative"),
 ("Backend error on load","Unstable load","Apply during error window","-","'ExpectedStartOfValueNotFound' surfaced; retry","Medium","Negative"),
 ("Transfer to same executive","No-op transfer","Transfer to current assignee","-","Handled; no false change","Low","Negative"),
 ("Filter by current assignee","Scope leads","Set Current Assignee filter","-","Only that executive's leads","Medium","Functional"),
]
for c in lt_cases: addc(M,P,LT_PRE,c)

# CRM Dashboard (curated)
M,P="CRM Dashboard","CRM Dashboard Page"; DB_PRE="User logged in; dashboard has data."
db_cases=[
 ("Dashboard loads","Widgets render","Open /crm-dashboard","-","All widgets load; no perpetual spinner","High","Functional"),
 ("Branch filter","Filter by branch","Select branch → Apply","-","All sections update for branch","High","Functional"),
 ("Executive filter","Filter by executive","Select executive → Apply","-","Data scoped to executive","High","Functional"),
 ("Period filter","Filter by period","Select This Week/Month/Year","-","Counts change per period","High","Functional"),
 ("Custom date range","Custom period","Pick from/to → Apply","-","Data within range","Medium","Functional"),
 ("Invalid custom range","From>To","From later than To","-","Blocked/validation","Medium","Negative"),
 ("Sum integrity","Totals reconcile","Compare badges vs exec table vs footer","-","Badge = Σ exec column = footer total","High","Data Integrity"),
 ("Drill-down count parity","Badge→list","Click a status badge (e.g. Warm)","-","Inner list count = badge count","High","Data Integrity"),
 ("Filter inheritance","Carry filters","Drill from filtered dashboard","-","Inner list inherits branch/exec/period","High","Functional"),
 ("Back retains filters","Navigation","Drill → Back","-","Dashboard keeps prior filters","Medium","Functional"),
 ("Win/Conversion/Lost %","Metric formula","Read rates vs counts","-","Ratios consistent with counts","Medium","Data Integrity"),
 ("Empty period","No data","Pick empty period","-","All zeros; no error","Medium","Negative"),
 ("Sales View toggle","View switch","Toggle Sales View","-","View/metrics recompute consistently","Medium","Functional"),
 ("Executive-wise table","Per-exec rows","Inspect table","-","Each executive row + footer total correct","Medium","Functional"),
 ("Slow/large load","Heavy data","This Year + All","-","Loads acceptably; sums still reconcile","Medium","Performance"),
]
for c in db_cases: addc(M,P,DB_PRE,c)

# Settings masters CRUD negatives (curated)
M,P="Settings","Lead Source / Lead Status / Item Category"; ST_PRE="User logged in; settings page open."
st_cases=[
 ("Create master","Add record","New → name → Save","unique name","Created; appears in list","High","Positive"),
 ("Duplicate master","Unique constraint","Create same name twice","existing name","'...already exist' rejection","High","Negative"),
 ("Empty name","Required validation","Save with blank name","-","Validation; not saved","High","Negative"),
 ("Whitespace name","Spaces only","Save '   '","'   '","Rejected/trimmed","Medium","Negative"),
 ("Max length","Boundary","Name > limit (e.g. 30)","31 chars","Truncated/blocked at limit","Medium","Boundary"),
 ("Special chars name","Symbols","Name '@#$%'","@#$%","Handled/validated","Low","Validation"),
 ("Edit master","Rename","Pencil → rename → Save","new name","Row updated","High","Positive"),
 ("Edit to duplicate","Rename clash","Rename to existing name","existing","Rejected (duplicate)","Medium","Negative"),
 ("Delete master","Remove","Trash → confirm","-","Row removed","High","Positive"),
 ("Delete in-use master","Referenced record","Delete a source used by leads","-","Blocked/warns about usage","Medium","Negative"),
 ("Cancel delete","Abort","Trash → cancel","-","Row retained","Low","Negative"),
 ("Search master","Filter list","Search by name","-","Matching rows shown","Medium","Functional"),
 ("Re-runnable create","No residue","Create→delete repeatedly","-","No duplicate buildup","Medium","Data Integrity"),
]
for c in st_cases: addc(M,P,ST_PRE,c)

# ============================================================ assign IDs
prefix_map={"Login":"LOGIN","CRM List":"LIST","New Enquiry":"ENQ","Save Enquiry":"SAVE",
 "Follow-up":"FU","Follow-up Reschedule":"FUR","Attachment":"ATT","Security":"SEC",
 "Performance":"PERF","Responsive & Compatibility":"RESP","Quotation":"QUO",
 "Inventory":"ITEM","Lead Transfer":"LT","CRM Dashboard":"DASH","Settings":"SET"}
counters={}
for r in rows:
    pfx=prefix_map.get(r[0],"TC"); counters[pfx]=counters.get(pfx,0)+1
    r[2]=f"TC-{pfx}-{counters[pfx]:03d}"

# ============================================================ build workbook
wb=Workbook()
HEAD_FILL=PatternFill("solid",fgColor="1F4E78"); HEAD_FONT=Font(bold=True,color="FFFFFF",name="Arial",size=10)
BASE_FONT=Font(name="Arial",size=10); WRAP=Alignment(wrap_text=True,vertical="top")
THIN=Side(style="thin",color="D9D9D9"); BORDER=Border(left=THIN,right=THIN,top=THIN,bottom=THIN)
PRIO_FILL={"High":PatternFill("solid",fgColor="F8CBAD"),"Medium":PatternFill("solid",fgColor="FFE699"),"Low":PatternFill("solid",fgColor="E2EFDA")}
WIDTHS=[14,26,16,30,40,28,46,26,40,10,16,14,28]

def style_sheet(ws, header, data, prio_col=None):
    ws.append(header)
    for j,_ in enumerate(header,1):
        c=ws.cell(1,j); c.fill=HEAD_FILL; c.font=HEAD_FONT; c.alignment=Alignment(wrap_text=True,vertical="center",horizontal="center"); c.border=BORDER
    for row in data:
        ws.append(row)
    for i in range(2,len(data)+2):
        for j in range(1,len(header)+1):
            c=ws.cell(i,j); c.font=BASE_FONT; c.alignment=WRAP; c.border=BORDER
        if prio_col:
            pc=ws.cell(i,prio_col); pc.fill=PRIO_FILL.get(pc.value,PatternFill())
    for j,w in enumerate(WIDTHS[:len(header)],1):
        ws.column_dimensions[get_column_letter(j)].width=w
    ws.freeze_panes="A2"
    ws.auto_filter.ref=f"A1:{get_column_letter(len(header))}{len(data)+1}"

ws=wb.active; ws.title="Test Cases"
style_sheet(ws, COLS, rows, prio_col=10)

# ---- Missed Scenario Review
mcols=["#","Area","Edge Case / Hidden Risk","Why It Matters","Suggested Action"]
missed=[
 ["Backend instability","Intermittent 'Oops Error Code' on save & 'ExpectedStartOfValueNotFound' on list load","Save/list randomly fail; data may not persist","Server-side fix; tests retry & surface error"],
 ["Tenant data variance","Dev tenant empty vs Test tenant seeded (customers/items/categories differ)","Data-dependent tests fail on fresh tenants","Make tests self-seed data"],
 ["Items delete absent on dev","/items list shows Edit only (no Delete) on some tenants","Cannot clean up created items; residue grows","Provide delete or admin cleanup"],
 ["Success message wording differs","'Saved Successfully' vs 'You're done!!' across tenants/pages","Automation success-detection breaks","Standardize toasts / classify broadly"],
 ["AJAX skeleton timing","Dashboard/overview load via spinners; reads before render","Flaky counts / clicks","Wait for content, not fixed timeouts"],
 ["Duplicate phone AND email","New customer modal rejects duplicate of either","Re-runs collide if data reused","Use unique name+phone+email per run"],
 ["Followup button disappears post-conversion","No follow-up after quotation conversion","Lifecycle order matters","Document; test order"],
 ["Search needs Enter to filter","List filter sometimes requires Enter/trigger","Row not found after create/edit","Retry filter; press Enter"],
 ["Concurrency / double submit","Rapid Save can create duplicates","Data integrity risk","Disable button on submit"],
 ["Session expiry mid-action","Long forms may expire before save","Data loss","Warn & preserve draft"],
 ["File upload security","exe/virus/oversized uploads","Security & storage risk","Server-side type+size+AV checks"],
 ["IDOR / direct URL","/enquiry-overview/{id} guessable","Unauthorized data access","Enforce per-record authorization"],
 ["Percent metric denominator","Win/Conversion/Lost rate base unclear","Wrong KPI interpretation","Confirm formula with product owner"],
 ["Pagination vs export 'all'","Export may export only current page","Incomplete exports","Verify export scope"],
 ["Time zone / locale","Dates/times across regions","Off-by-day follow-ups","Standardize TZ handling"],
]
ws2=wb.create_sheet("Missed Scenario Review")
mdata=[[i+1,*m] for i,m in enumerate(missed)]
style_sheet(ws2, mcols, mdata)
for j,w in enumerate([6,26,46,40,40],1): ws2.column_dimensions[get_column_letter(j)].width=w

# ---- Regression Suite (subset: High priority functional)
ws3=wb.create_sheet("Regression Suite")
reg=[r for r in rows if r[9]=="High" and r[10] in ("Positive","Functional","Negative","Data Integrity")][:60]
style_sheet(ws3, COLS, reg, prio_col=10)

# ---- Smoke Test Cases (critical release validation)
smoke_titles=["Valid login","Grid data load","Save complete data","Create follow-up",
 "Convert","Transfer to executive","Create lead source","Excel export"]
smoke=[
 ["CRM","Login Page","TC-SMK-001","Login","Login with valid credentials","Valid tenant","Open login → enter creds → Login","Lesol_dev/admin/123","Reaches /home","High","Smoke","Not Executed","Release gate"],
 ["CRM","List Page","TC-SMK-002","List loads","Leads list renders","Logged in","Open /leads","-","Rows + tabs load","High","Smoke","Not Executed",""],
 ["CRM","New Enquiry","TC-SMK-003","Create enquiry","Create lead with item","Logged in","New Enquiry → fill → add item → Save","valid","Saved; overview opens","High","Smoke","Not Executed",""],
 ["CRM","Overview","TC-SMK-004","Add follow-up","Record a follow-up","Enquiry exists","Followup → status → Save","Interested","Follow-up saved","High","Smoke","Not Executed",""],
 ["CRM","Overview","TC-SMK-005","Convert to quotation","Enquiry → quotation","Enquiry exists","Create Quotation → Save","-","Quotation created","High","Smoke","Not Executed",""],
 ["CRM","Lead Transfer","TC-SMK-006","Transfer lead","Reassign to executive","Leads exist","Apply → select → executive → Transfer","-","Assignee updated","High","Smoke","Not Executed",""],
 ["CRM","Settings","TC-SMK-007","Create lead source","Master create","Logged in","New Lead Source → name → Save","unique","Created & listed","High","Smoke","Not Executed",""],
 ["Inventory","Items","TC-SMK-008","Create item","Product create","Logged in","New Item → name+category → Save","unique","Item created","High","Smoke","Not Executed",""],
 ["CRM","Dashboard","TC-SMK-009","Dashboard loads","Widgets + counts","Logged in","Open CRM dashboard","-","Widgets render; totals reconcile","High","Smoke","Not Executed",""],
 ["CRM","List Page","TC-SMK-010","Excel export","Export data","List has rows","Click Export Data","-","Valid xlsx downloaded","High","Smoke","Not Executed",""],
]
ws4=wb.create_sheet("Smoke Test Cases")
style_sheet(ws4, COLS, smoke, prio_col=10)

# ---- Summary on first sheet position? add a cover note via sheet order
wb.move_sheet("Test Cases", -(len(wb.sheetnames)-1))
wb.save("CRM_Test_Cases.xlsx")
print("TOTAL test cases (main sheet):", len(rows))
print("Sheets:", wb.sheetnames)
