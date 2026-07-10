# -*- coding: utf-8 -*-
"""Generate two Excel docs of the AUTOMATED scenarios + step procedures:
   CRM_Automation_Scenarios.xlsx  and  TaskManagement_Automation_Scenarios.xlsx"""
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import os

REPO = os.path.join(os.path.dirname(__file__), "..")
DL = r"C:\Users\PROBOOK\Downloads"

HDR = ["Module", "Test Case ID", "Scenario", "Automation Steps (Procedure)", "Expected Result", "Spec File", "Status"]

# ── CRM rows: (Module, ID, Scenario, Steps, Expected, Spec, Status) ──
S_LOGIN = "erp/crm/tests/crm_login.spec.js"
S_HOME  = "erp/crm/tests/crm_homepage.spec.js"
S_ITEM  = "erp/crm/tests/crm_item.spec.js"
S_ENQ   = "erp/crm/tests/crm_enquiry.spec.js"
S_FU    = "erp/crm/tests/crm_followup.spec.js"
S_QT    = "erp/crm/tests/crm_quotation.spec.js"
S_FLOW  = "erp/crm/tests/enquiry_flow.spec.js"

crm = [
 ("Login","Login_01","Login page loads","1. Open /login\n2. Wait for the JS-rendered form","Company Code, Username, Password fields are visible","",S_LOGIN,"Pass"),
 ("Login","Login_02","Login with valid credentials","1. Enter Company/Username/Password\n2. Click Sign In","User is logged in; URL leaves /login",S_LOGIN,"Pass"),
 ("Login","Login_03","Mapped company login","1. Enter valid mapped company + credentials\n2. Sign In","Login succeeds",S_LOGIN,"Pass"),
 ("Login","Login_04","Login using Enter key","1. Fill credentials\n2. Press Enter in the password field","User is logged in without clicking the button",S_LOGIN,"Pass"),
 ("Login","Login_05","Password visibility (eye) icon","1. Type a password\n2. Click the eye/.password-toggle icon","Field type flips password->text (password revealed)",S_LOGIN,"Pass"),
 ("Login","Login_06","Remember Password option","1. Look for a Remember-Password checkbox","No such checkbox on this build",S_LOGIN,"Skip"),
 ("Login","Login_07","Forgot Password link","1. Locate 'Forgot password ?' link\n2. Click it","Link is present and clickable",S_LOGIN,"Pass"),
 ("Login","Login_08","Redirect to dashboard","1. Login with valid credentials","Lands on /home dashboard",S_LOGIN,"Pass"),

 ("Homepage","Home_01,02,24","Homepage loads + welcome + Create New","1. Login -> /home\n2. Read greeting + Create New","Greeting 'Hey, <user>' shown; Create New visible",S_HOME,"Pass"),
 ("Homepage","Home_03,04","New Leads & lead status counts","1. Open /leads\n2. Read #tab-lead-new/infollowup/won/lost counts\n3. Click New Leads tab","Lead status tabs + counts shown; tab opens the list",S_HOME,"Pass"),
 ("Homepage","Home_05,06,07,08","Followups Today/Overdue counts + drill-down","1. Open /followups\n2. Read Today & Delayed(overdue) counts\n3. Click Overdue tab","Today/Overdue counts shown; drill-down works",S_HOME,"Pass"),
 ("Homepage","Home_09,10","Won/Completed leads listing","1. Open /leads\n2. Read Won count\n3. Click Won tab","Won/Completed leads listing reachable",S_HOME,"Pass"),
 ("Homepage","Home_11,12","Today's Schedule section","1. Open /home\n2. Check Today's Schedule section","Today's Schedule section present",S_HOME,"Pass"),
 ("Homepage","Home_13,14","Follow-up History","1. Open /followups\n2. Check history/listing of past follow-ups","Follow-up history/listing present",S_HOME,"Pass"),
 ("Homepage","Home_15,16","Summary lead classification","1. Open /crm-dashboard\n2. Read Summary words","Total/New/Cold/Warm/Hot/Won/Lost classification shown",S_HOME,"Pass"),
 ("Homepage","Home_17,18,19","Executive filter on CRM Dashboard","1. Open /crm-dashboard\n2. Open filters\n3. Look for Executive/Added-By filter","Executive/Added-By filter available for admin",S_HOME,"Pass"),
 ("Homepage","Home_20,21,22,23","Timeline & Calendar icons","1. Open /home\n2. Check for calendar + clock/timeline icons","Timeline/Calendar entry points present",S_HOME,"Pass"),
 ("Homepage","Home_24,25,26","Create New -> Enquiry & Quotation","1. Open Create New menu\n2. Read options","Create New exposes Task/Enquiry/Quotation",S_HOME,"Pass"),

 ("Item","Item_01","Create item (valid mandatory)","1. Open /item\n2. Enter Item Name + Variant\n3. Save","Item created and listed",S_ITEM,"Pass"),
 ("Item","Item_02","Create item (multiple fields)","1. Enter Name + Variant + Description\n2. Save","Item saved with details",S_ITEM,"Pass"),
 ("Item","Item_03","Access Add Item without login","1. Clear session\n2. Navigate to /item","Redirected to Login / form not shown",S_ITEM,"Pass"),
 ("Item","Item_04","Session timeout while creating","(Not deterministically automatable)","Redirect/expiry on timeout",S_ITEM,"Skip"),
 ("Item","Item_05","Create without Item Name","1. Leave Item Name blank\n2. Save","Validation 'Please provide a name for the item'",S_ITEM,"Pass"),
 ("Item","Item_06","Whitespace-only Item Name","1. Enter spaces as name\n2. Save","Required/invalid validation shown",S_ITEM,"Pass"),
 ("Item","Item_07,08,09","Price validations","No Price field on this build's item form","N/A",S_ITEM,"Skip"),
 ("Item","Item_10","Duplicate Item Name","1. Create item\n2. Create again with same name","'Item name already exist' rejection",S_ITEM,"Pass"),
 ("Item","Item_11","Save creates and lists","1. Create valid item\n2. Search list","Item persisted to the grid",S_ITEM,"Pass"),
 ("Item","Item_12","Search Item","1. Search created name\n2. Search a random name","Matching item found; random not found",S_ITEM,"Pass"),
 ("Item","Item_13","Edit Item","1. Find row -> Edit\n2. Change name -> Save","Updated name appears in list",S_ITEM,"Pass"),
 ("Item","Item_14","Delete Item","/items has no Delete control on this tenant","Edit-only tenant",S_ITEM,"Skip"),
 ("Item","Item_15","Cancel while adding","1. Enter data\n2. Click Cancel","Returns without saving",S_ITEM,"Pass"),

 ("Enquiry","ENQ-01","Access Enquiry from Create New","1. Create New -> Enquiry (/enquiry)","Add Enquiry form reachable",S_ENQ,"Pass"),
 ("Enquiry","ENQ-02,03,04,15","Add Enquiry form fields","1. Open form\n2. Read Branch options, auto Date, auto ENQ number, Lead Source options","Branch=Kannur/Kasargod, date auto, ENQ-### auto, sources listed",S_ENQ,"Pass"),
 ("Enquiry","ENQ-05","Customer search picker","1. Use the phone/customer search icon\n2. Pick a customer","Customer name auto-fills",S_ENQ,"Pass"),
 ("Enquiry","ENQ-08","Followup Status options","1. Read #followup dropdown options","Interested/Got the business/Not interested etc. present",S_ENQ,"Pass"),
 ("Enquiry","ENQ-09,12","Lead Quality conditional","1. Select status 'New' -> LQ hidden\n2. Select 'Interested' -> LQ shown (Cold/Warm/Hot)","Lead Quality only for In-Followup",S_ENQ,"Pass"),
 ("Enquiry","ENQ-10,11","Description for Won/Lost","1. Select 'Got the business' and 'Not interested'","Description field visible for Won and Lost",S_ENQ,"Pass"),
 ("Enquiry","ENQ-13","Next Followup Date","1. Fill #next-followup-date","Date accepted",S_ENQ,"Pass"),
 ("Enquiry","ENQ-16,18","Item selection + multiple items","1. Add item 'Inverter'\n2. Add item 'Generator'","Item rows added under Enquired For",S_ENQ,"Pass"),
 ("Enquiry","ENQ-19,21","Save enquiry -> Overview","1. Fill mandatory fields (+ new customer)\n2. Save","Saved; redirected to Enquiry Overview",S_ENQ,"Pass"),
 ("Enquiry","ENQ-20","Cancel returns to listing","1. Click Cancel","Returns to Leads listing without saving",S_ENQ,"Pass"),
 ("Enquiry","ENQ-22,23,24","Enquiry Overview details","1. Open an enquiry overview","Customer/status/details displayed",S_ENQ,"Pass"),
 ("Enquiry","ENQ-25,26,27","Overview actions (View/Edit/Followup)","1. Open enquiry overview\n2. Check available actions","Overview opens with Followup/Edit actions",S_ENQ,"Pass"),
 ("Enquiry","ENQ-28","Create Quotation from enquiry","1. Create fresh enquiry\n2. Click Create Quotation","Quotation flow reached",S_ENQ,"Pass"),
 ("Enquiry","ENQ-07","Assign To dropdown","No visible Assign-To field (auto-assigned)","Verified via successful create",S_ENQ,"Skip"),
 ("Enquiry","ENQ-14","Business Value input","Captured in the follow-up modal on this build","Covered by Followup ENQ-35",S_ENQ,"Skip"),

 ("Followup","ENQ-29..35","Followup popup + date + conditional fields","1. Seed enquiry -> open #followupModal\n2. Check date auto-fill, status options\n3. Interested->LQ shown; Won/Lost->only Description\n4. Business Value editable","Modal structure + conditional fields verified",S_FU,"Pass"),
 ("Followup","ENQ-36,38","Save follow-up -> history","1. Fill follow-up\n2. Save\n3. Open history","Follow-up saved and visible in history",S_FU,"Pass"),
 ("Followup","ENQ-37","Cancel closes popup","1. Open modal\n2. Click Cancel/Close","Popup closes without saving",S_FU,"Pass"),
 ("Followup","ENQ-39..42","Latest follow-up editable/deletable","1. Add a follow-up\n2. Check latest row controls","Latest exposes edit/delete; older read-only by design",S_FU,"Pass"),
 ("Followup","QT-019..028","Quotation follow-up (same modal)","Reuses the identical #followupModal (ENQ-29..42)","No separate behaviour",S_FU,"Skip"),

 ("Quotation","QT-003,004,006,010,011","Auto-fill from enquiry + items + totals","1. Enquiry -> Create Quotation\n2. Read prefilled fields, item rows, totals\n3. Edit a field","All auto-filled except 'Valid Upto'; items+totals present; editable",S_QT,"Pass"),
 ("Quotation","QT-007","Terms & Conditions editable","1. Type into #terms-and-condition","Terms accepts input",S_QT,"Pass"),
 ("Quotation","QT-008,012,013,014,015,016,017,018","Save quotation -> Overview + actions","1. Set Valid Upto\n2. Click Save (#btn-save-quotation)","Saved to Overview with actions",S_QT,"Pass"),
 ("Quotation","QT-001,002,009","Create New -> Quotation page","1. Create New -> Quotation","Quotation creation page reachable",S_QT,"Pass"),

 ("CRM Flow","TC-01","Login with valid credentials","1. Login","Logged in",S_FLOW,"Pass"),
 ("CRM Flow","TC-02","Create new enquiry (new customer)","1. Fill form + new customer + item\n2. Save","Enquiry created",S_FLOW,"Pass"),
 ("CRM Flow","TC-02B","Create enquiry (existing customer)","1. Search & choose existing customer\n2. Save","Enquiry created for existing customer",S_FLOW,"Pass"),
 ("CRM Flow","TC-03","Open created enquiry","1. Open enquiry from list","Overview opens",S_FLOW,"Pass"),
 ("CRM Flow","TC-04","Add a follow-up","1. Open Followup modal\n2. Save","Follow-up added",S_FLOW,"Pass"),
 ("CRM Flow","TC-05","Follow-up visible in listing","1. Open follow-up history","Follow-up listed",S_FLOW,"Pass"),
 ("CRM Flow","TC-06","Convert enquiry to quotation","1. Create Quotation from overview","Quotation created",S_FLOW,"Pass"),
 ("CRM Flow","TC-07","Quotation appears in listing","1. Check quotation listing","Quotation listed",S_FLOW,"Pass"),
 ("CRM Flow","TC-08","Update status In Follow-up","1. Follow-up status = Interested","Status -> In Follow-up",S_FLOW,"Pass"),
 ("CRM Flow","TC-09","Update status Won","1. Follow-up status = Got the business","Status -> Won",S_FLOW,"Pass"),
 ("CRM Flow","TC-10","Update status Lost","1. Follow-up status = Not interested","Status -> Lost",S_FLOW,"Pass"),
 ("CRM Flow","TC-11","Records visible in listing","1. Open enquiry listing","Records visible",S_FLOW,"Pass"),
 ("CRM Flow","TC-12","Lead Transfer","1. Filter leads\n2. Transfer to executive\n3. Verify assignee","Assignee changes",S_FLOW,"Pass"),
 ("CRM Flow","TC-13","Lead Sources (Settings)","1. Create a lead source","Source listed",S_FLOW,"Pass"),
 ("CRM Flow","TC-14","Lead Status (Settings)","1. Create followup status with Nature","Status created",S_FLOW,"Pass"),
 ("CRM Flow","TC-15","Item Categories","1. Create category\n2. Retry duplicate","Duplicate rejected",S_FLOW,"Pass"),
 ("CRM Flow","TC-16","Items create/duplicate/delete","1. Create Product\n2. Duplicate rejected\n3. Delete","Full CRUD verified",S_FLOW,"Pass"),
]

# ── Task Management rows ──
T_BASE = "erp/task-management/tests/task_management.spec.js"
T_MOD  = "erp/task-management/tests/task_management_modal.spec.js"
T_DET  = "erp/task-management/tests/task_management_details.spec.js"
T_MU   = "erp/task-management/tests/task_management_multiuser.spec.js"

tm = [
 ("My Tasks","TM-01","My Tasks page loads with status tabs","1. Open /my-tasks\n2. Read status tabs","Today/Delayed/Upcoming/Unscheduled/Completed tabs present",T_BASE,"Pass"),
 ("Create","TM-02","Create a task and verify listed","1. /task -> fill -> Save\n2. Search My Tasks","Task created and findable",T_BASE,"Pass"),
 ("Navigation","TM-03","All Task-Management pages reachable","1. Visit My Tasks/Delegated/To-Do/Calendar/Daily Report","All pages load",T_BASE,"Pass"),
 ("Create","TM-04","Create scheduled task (Task for Later)","1. Later mode + date/time\n2. Save","Scheduled task saved",T_BASE,"Pass"),
 ("Negative","TM-05","Task without Task Type rejected","1. Leave Task Type = Choose\n2. Save","'Please choose valid task type'",T_BASE,"Pass"),
 ("Create","TM-06","Create Online Meeting, High priority","1. Type=Online Meeting, Priority=High\n2. Save","Task created",T_BASE,"Pass"),
 ("My Tasks","TM-07","Status tabs navigable","1. Click each My Tasks tab","All tabs navigable",T_BASE,"Pass"),
 ("Report","TM-08","Daily Activity Report loads","1. Open /daily-activity-report","Report loads with rows",T_BASE,"Pass"),

 ("Modal","TM-09","Create New -> Task modal structure","1. Create New -> Task\n2. Assert Branch/Task Type/Priority options, title, party/participant controls","Add Task modal matches documented spec (TC_001-016)",T_MOD,"Pass"),
 ("Create","TM-10","Create Instant Task via modal (Scenario 1)","1. Open modal\n2. Fill Type/Priority/Task\n3. Save","Instant task created; appears in My Tasks",T_MOD,"Pass"),
 ("Create","TM-11","Task for Later - Hosts + scheduling (Scenario 2)","1. Later mode\n2. Host field + deadline toggle + date/time\n3. Save","Scheduled task created with Host",T_MOD,"Pass"),
 ("Create","TM-12","Repeat mode - recurring + create (NEW)","1. Repeat mode\n2. Recurrence Daily/Weekly/Monthly + Start/End Time + From/To Date\n3. Save","Recurring Daily task created",T_MOD,"Pass"),
 ("Negative","TM-13","Save without Task Type rejected","1. Skip Task Type\n2. Save","'Please choose valid task type'",T_MOD,"Pass"),
 ("Negative","TM-14","Save without Task title rejected","1. Leave title empty\n2. Save","'Please add task'",T_MOD,"Pass"),
 ("Pages","TM-15","My Task page columns (TC_054-055)","1. Open /my-tasks\n2. Read columns","Status/Task/Task Type/Added on columns present",T_MOD,"Pass"),
 ("Pages","TM-16","Created Task page + actions (TC_048-053)","1. Open /created-tasks\n2. Read columns + row actions","Page reachable with actions",T_MOD,"Pass"),
 ("Pages","TM-17","Delegated Tasks - Assignees column","1. Open /delegated-tasks","Assignees column present",T_MOD,"Pass"),
 ("Pages","TM-18","Unscheduled Task page + row actions (Scenario 11)","1. Open /unscheduled-tasks\n2. Check edit/delete/overview actions","Page + row actions present",T_MOD,"Pass"),
 ("My Tasks","TM-19","Status tabs Pending/Overdue/Completed (TC_057-059)","1. Navigate each status tab","Tabs navigable",T_MOD,"Pass"),
 ("Lifecycle","TM-20","Lifecycle controls on dashboard (Scenario 4)","1. Open /home\n2. Check start/hold/end controls + sections","Controls + Running/On-Hold sections present",T_MOD,"Pass"),
 ("Report","TM-21","Daily Activity Report (TC_069-072, Scenario 12)","1. Open Daily Activity Report","Loads with rows",T_MOD,"Pass"),
 ("Views","TM-22","Calendar & Timeline reachable (Scenario 6)","1. Open Calendar\n2. Open Timeline","Both reachable",T_MOD,"Pass"),
 ("Multi-user","TM-23","Participant/admin visibility placeholder","Covered by MU-01..03 (2nd login)","Placeholder for traceability",T_MOD,"Skip"),

 ("Notes/Docs","TM-24","Add note + upload document (Scenario 7)","1. Open a task's Details panel\n2. Type note in #txtChat (Enter)\n3. Attach a document (#file-input-document)","Note in activity log + document uploaded",T_DET,"Pass"),
 ("Edit","TM-25","Edit an existing task title (Scenario 13)","1. Details -> Menu -> Edit Task\n2. Change title -> Save","Updated title persists in My Tasks",T_DET,"Pass"),
 ("Reschedule","TM-26","Reschedule a task (Scenario 14)","1. Details -> Menu -> Reschedule\n2. New date/time -> Save","Reschedule saved",T_DET,"Pass"),
 ("Lead","TM-27","Add a Lead from a task (Scenario 15)","1. Details -> Menu -> Add Lead","Opens /enquiry lead form",T_DET,"Pass"),
 ("Lifecycle","TM-28","Lifecycle Hold -> Resume -> End (Scenario 4)","1. Create running task\n2. Hold (confirm time) -> row=Hold\n3. Resume\n4. End (confirm)","Hold verified via row status; Resume/End execute",T_DET,"Pass"),

 ("Multi-user","MU-01","Host-assigned task visible to assignee (Scenario 10)","1. Admin assigns current-date Task-for-Later Host=HAFNEETHA\n2. HAFNEETHA logs in -> My Tasks","Assignee sees the task",T_MU,"Pass"),
 ("Multi-user","MU-02","Participant sees Instant task (Scenario 5)","1. Admin adds HAFNEETHA as Participant on Instant task\n2. HAFNEETHA logs in","Task visible under Today",T_MU,"Pass"),
 ("Multi-user","MU-03","Admin Calendar/Timeline oversight (Scenario 6)","1. Admin opens Calendar + Timeline","Both reachable for oversight",T_MU,"Pass"),
]

def build(path, title, rows):
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = title
    thin = Side(style="thin", color="D0D0D0"); border = Border(left=thin,right=thin,top=thin,bottom=thin)
    hf = PatternFill("solid", fgColor="1F3864")
    for c,h in enumerate(HDR, start=1):
        cell = ws.cell(1,c,h); cell.fill=hf; cell.font=Font(bold=True,color="FFFFFF"); cell.alignment=Alignment(horizontal="center",vertical="center"); cell.border=border
    stf={"Pass":PatternFill("solid",fgColor="C6EFCE"),"Skip":PatternFill("solid",fgColor="D9D9D9")}
    stc={"Pass":"006100","Skip":"595959"}
    for r,row in enumerate(rows, start=2):
        for c,val in enumerate(row, start=1):
            cell=ws.cell(r,c,val); cell.alignment=Alignment(vertical="top",wrap_text=True); cell.border=border
        s=row[6]; sc=ws.cell(r,7); sc.fill=stf.get(s,PatternFill()); sc.font=Font(bold=True,color=stc.get(s,"000000")); sc.alignment=Alignment(horizontal="center",vertical="top")
    widths=[14,18,34,50,40,42,9]
    for i,w in enumerate(widths, start=1): ws.column_dimensions[chr(64+i)].width=w
    ws.freeze_panes="A2"; ws.auto_filter.ref=f"A1:G{len(rows)+1}"
    wb.save(path)
    return len(rows)

for fn, title, rows in [("CRM_Automation_Scenarios.xlsx","CRM Automation Scenarios",crm),
                        ("TaskManagement_Automation_Scenarios.xlsx","Task Mgmt Automation Scenarios",tm)]:
    n1=build(os.path.join(DL,fn), title, rows)
    build(os.path.join(REPO,fn), title, rows)
    print(f"{fn}: {n1} rows -> Downloads + repo")
