# -*- coding: utf-8 -*-
"""Add an automation Status column to the CRM Test Case Excel, based on the
actual Playwright runs in this repo. Honest 3-state: Pass / Partial / Skip."""
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
import os

SRC = r"C:\Users\PROBOOK\Downloads\CRM- Test Case.xlsx"
OUT = r"C:\Users\PROBOOK\Downloads\CRM- Test Case - Automation Status.xlsx"
REPO = os.path.join(os.path.dirname(__file__), "..", "docs", "excel", "CRM- Test Case - Automation Status.xlsx")

# Pass = directly asserted & green | Partial = covered by a passing test but the
# deep/dynamic behaviour not separately asserted | Skip = N/A on this build (feature/role absent)
STATUS = {
  # Login (crm_login.spec.js)
  "Login_01":"Pass","Login_02":"Pass","Login_03":"Pass","Login_04":"Pass","Login_05":"Pass",
  "Login_06":"Skip","Login_07":"Pass","Login_08":"Pass",
  # Homepage (crm_homepage.spec.js)
  "Home_01":"Pass","Home_02":"Pass","Home_03":"Pass","Home_04":"Pass","Home_05":"Pass","Home_06":"Pass",
  "Home_07":"Pass","Home_08":"Pass","Home_09":"Pass","Home_10":"Pass","Home_11":"Pass","Home_12":"Partial",
  "Home_13":"Pass","Home_14":"Partial","Home_15":"Pass","Home_16":"Partial","Home_17":"Pass","Home_18":"Partial",
  "Home_19":"Skip","Home_20":"Pass","Home_21":"Partial","Home_22":"Pass","Home_23":"Partial","Home_24":"Pass",
  "Home_25":"Pass","Home_26":"Pass",
  # Item (crm_item.spec.js)
  "Item_01":"Pass","Item_02":"Pass","Item_03":"Pass","Item_04":"Skip","Item_05":"Pass","Item_06":"Pass",
  "Item_07":"Skip","Item_08":"Skip","Item_09":"Skip","Item_10":"Pass","Item_11":"Pass","Item_12":"Pass",
  "Item_13":"Pass","Item_14":"Skip","Item_15":"Pass",
  # Enquiry (crm_enquiry.spec.js)
  "ENQ-01":"Pass","ENQ-02":"Pass","ENQ-03":"Pass","ENQ-04":"Pass","ENQ-05":"Pass","ENQ-06":"Pass",
  "ENQ-07":"Skip","ENQ-08":"Pass","ENQ-09":"Pass","ENQ-10":"Pass","ENQ-11":"Pass","ENQ-12":"Pass",
  "ENQ-13":"Pass","ENQ-14":"Skip","ENQ-15":"Pass","ENQ-16":"Pass","ENQ-17":"Partial","ENQ-18":"Pass",
  "ENQ-19":"Pass","ENQ-20":"Pass","ENQ-21":"Pass","ENQ-22":"Pass","ENQ-23":"Pass","ENQ-24":"Partial",
  "ENQ-25":"Pass","ENQ-26":"Partial","ENQ-27":"Partial","ENQ-28":"Pass",
  # Followup (crm_followup.spec.js)
  "ENQ-29":"Pass","ENQ-30":"Pass","ENQ-31":"Pass","ENQ-32":"Pass","ENQ-33":"Pass","ENQ-34":"Pass",
  "ENQ-35":"Pass","ENQ-36":"Pass","ENQ-37":"Pass","ENQ-38":"Pass","ENQ-39":"Pass","ENQ-40":"Partial",
  "ENQ-41":"Pass","ENQ-42":"Partial",
  # Quotation (crm_quotation.spec.js)
  "QT-001":"Pass","QT-002":"Pass","QT-003":"Pass","QT-004":"Pass","QT-005":"Partial","QT-006":"Pass",
  "QT-007":"Pass","QT-008":"Pass","QT-009":"Partial","QT-010":"Pass","QT-011":"Pass","QT-012":"Pass",
  "QT-013":"Pass","QT-014":"Partial","QT-015":"Partial","QT-016":"Partial","QT-017":"Partial","QT-018":"Partial",
  # Quotation follow-up (same #followupModal as ENQ-29..42)
  "QT-019":"Pass","QT-020":"Pass","QT-021":"Pass","QT-022":"Pass","QT-023":"Pass","QT-024":"Pass",
  "QT-025":"Pass","QT-026":"Pass","QT-027":"Pass","QT-028":"Pass",
}
REMARK = {
  "Login_06":"No Remember-Password checkbox on this build",
  "Home_12":"Today's Schedule present; exact match-to-Followups not asserted",
  "Home_14":"History present; dynamic add-to-history not asserted",
  "Home_16":"Summary present; filter-driven recalculation not asserted",
  "Home_18":"Exec filter present; data refresh on selection not asserted",
  "Home_19":"Needs a limited/non-admin role login (hidden-for-limited)",
  "Home_21":"Timeline entry-point present; timeline content not asserted",
  "Home_23":"Calendar entry-point present; calendar content not asserted",
  "Item_04":"Session timeout not deterministically automatable",
  "Item_07":"No Price field on this build's item form",
  "Item_08":"No Price field on this build's item form",
  "Item_09":"No Price field on this build's item form",
  "Item_14":"/items exposes no Delete control (Edit-only) on this tenant",
  "ENQ-07":"No Assign-To field on this build (auto-assigned; create succeeds)",
  "ENQ-14":"Business Value lives in the follow-up modal (verified by ENQ-35)",
  "ENQ-17":"Item selection verified; inline new-item creation not asserted",
  "ENQ-24":"Overview loads; item/quantity row not separately asserted",
  "ENQ-26":"Delete-after-followup rule covered by design (latest-only)",
  "ENQ-27":"View/overview verified; full edit round-trip not asserted",
  "ENQ-40":"Older follow-up read-only by design (latest-row check)",
  "ENQ-42":"Older follow-up not deletable by design (latest-row check)",
  "QT-005":"Totals render; discount-inclusive-tax recalculation not asserted",
  "QT-009":"Cancel not separately asserted (mandatory Valid Upto enforced on save)",
  "QT-014":"Overview reachable; full edit round-trip not asserted",
  "QT-015":"Overview actions present (loose text assertion)",
  "QT-016":"Overview actions present (loose text assertion)",
  "QT-017":"Overview actions present (loose text assertion)",
  "QT-018":"Overview actions present (loose text assertion)",
}
def spec_for(tid):
  if tid.startswith("Login_"): return "tests/crm_login.spec.js"
  if tid.startswith("Home_"):  return "tests/crm_homepage.spec.js"
  if tid.startswith("Item_"):  return "tests/crm_item.spec.js"
  if tid.startswith("ENQ-"):
    n = int(tid.split("-")[1]); return "tests/crm_enquiry.spec.js" if n <= 28 else "tests/crm_followup.spec.js"
  if tid.startswith("QT-"):
    n = int(tid.split("-")[1]); return "tests/crm_followup.spec.js" if 19 <= n <= 28 else "tests/crm_quotation.spec.js"
  return ""

FILL = {"Pass":PatternFill("solid",fgColor="C6EFCE"),"Partial":PatternFill("solid",fgColor="FFEB9C"),
        "Skip":PatternFill("solid",fgColor="D9D9D9")}
FONTC = {"Pass":"006100","Partial":"9C6500","Skip":"595959"}

wb = openpyxl.load_workbook(SRC)
ws = wb["Demo Flow"]
# header cells for the 3 new columns (G,H,I)
hdr = PatternFill("solid",fgColor="1F3864")
for col, name in [(7,"Automation Status"),(8,"Automated By"),(9,"Remarks")]:
    c = ws.cell(row=1, column=col, value=name)
    c.fill = hdr; c.font = Font(bold=True, color="FFFFFF"); c.alignment = Alignment(horizontal="center")
ws.column_dimensions["G"].width = 18; ws.column_dimensions["H"].width = 30; ws.column_dimensions["I"].width = 52

counts = {"Pass":0,"Partial":0,"Skip":0,"?":0}
for r in range(2, ws.max_row+1):
    tid = ws.cell(row=r, column=2).value
    if not tid: continue
    tid = str(tid).strip()
    st = STATUS.get(tid)
    if not st:
        counts["?"] += 1; continue
    counts[st] += 1
    g = ws.cell(row=r, column=7, value=st)
    g.fill = FILL[st]; g.font = Font(bold=True, color=FONTC[st]); g.alignment = Alignment(horizontal="center")
    ws.cell(row=r, column=8, value=spec_for(tid))
    ws.cell(row=r, column=9, value=REMARK.get(tid, ""))

# Summary sheet
s = wb.create_sheet("Automation Summary")
s["A1"]="CRM Automation — Status Summary"; s["A1"].font=Font(bold=True,size=14)
rows=[("Status","Count"),("Pass",counts["Pass"]),("Partial",counts["Partial"]),
      ("Skip (N/A on this build)",counts["Skip"]),("Total mapped",counts["Pass"]+counts["Partial"]+counts["Skip"])]
for i,(a,b) in enumerate(rows, start=3):
    s.cell(row=i,column=1,value=a).font=Font(bold=(i==3)); s.cell(row=i,column=2,value=b).font=Font(bold=(i==3))
s["A9"]="Legend: Pass = asserted & green · Partial = covered by a passing test, deep/dynamic behaviour not separately asserted · Skip = feature/role absent on this build"
s.column_dimensions["A"].width=42; s.column_dimensions["B"].width=12

wb.save(OUT)
wb.save(REPO)
print("Counts:", counts)
print("Saved:", OUT)
print("Saved:", REPO)
