import openpyxl
wb = openpyxl.load_workbook(r"C:\Users\PROBOOK\Downloads\CRM- Test Case.xlsx", data_only=True)
print("SHEETS:", wb.sheetnames)
for ws in wb.worksheets:
    print(f"\n===== SHEET: {ws.title} (rows={ws.max_row}, cols={ws.max_column}) =====")
    # print header row
    rows = list(ws.iter_rows(values_only=True))
    if not rows: 
        print("(empty)"); continue
    hdr = rows[0]
    print("HDR:", [str(c).strip() if c is not None else '' for c in hdr])
