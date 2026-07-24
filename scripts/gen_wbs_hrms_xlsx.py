# -*- coding: utf-8 -*-
"""HRMS WBS tracker: Project -> Module -> Sub-module -> page-wise functionality
checks, with Estimated/Actual hours, daily log and a formula dashboard.

Same shape as gen_wbs_xlsx.py (CRM/Task Management) so both trackers behave
identically; the dashboard rolls up by SUB-MODULE here because every row shares
Module = "HRMS".

Estimates are grounded in the live-app study in hrms/docs/ and in the actual
Playwright automation of all 80 pages, so known effort drivers are priced in:
offcanvas-resident filters, SweetAlert2 confirm/validation dialogs, routed
create pages (vs modals), kanban/calendar/map/org-chart views, filter-first
report pages that render empty until queried, and data-dependent empty states.
"""
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
import os

REPO = os.path.join(os.path.dirname(__file__), "..", "docs", "excel", "WBS_HRMS.xlsx")

P = "Progbiz ERP"
M = "HRMS"
# (Sub-Module, Page / Functionality Check, Scope of check, Est h, Rationale)
ROWS = [
 # ─────────────────────────── CORE HR (18 pages) ───────────────────────────
 ("Core HR","Employee listing (/employees)","List load, Filter panel, incl-archived toggle, columns (Code/Name/Dept/Designation/Status), row actions, drill-down",3.0,"List + filter matrix; grid starts empty until filtered"),
 ("Core HR","Employee create / edit form","New Employee form: personal, branch/department/designation, contact, joining, mandatory-field negatives, save/cancel",4.0,"Largest master form in the module; full field matrix + negative paths"),
 ("Core HR","Sections (/sections)","Department-linked section create, inline form + grid, edit, delete, duplicate rejection",1.5,"Simple inline-form CRUD master"),
 ("Core HR","Worker Directory (/worker-directory)","Cards view, Org Chart view, department filter, name/designation search",2.5,"Two non-grid visualisations; org chart needs hierarchy data"),
 ("Core HR","Salary Revisions (/salary-revisions)","Raise Revision form (employee, effective date, new salary, reason), revision history grid, %-change, approval status",3.0,"Form + history + routes into the approval chain"),
 ("Core HR","Employee Salary Process (/employee-salary-process)","Branch selection, staff list, Basic/leave/payable computation, narration fields, save",3.5,"Arithmetic verification per employee; consumes attendance + deductions"),
 ("Core HR","Employee Deductions (/employee-deduction)","Branch/staff/date/deduction-type typeahead/amount/details, save, Cancel discards",2.5,"Duplicate element ids (#employee x2) + swal2 validation dialog"),
 ("Core HR","Employee Remarks (/employee-remark)","Reduced form (branch/employee/date/details), save, cancel",1.5,"Small form; header falsely reads 'Employee Deduction' (build bug to log)"),
 ("Core HR","Probation Dashboard (/hrms/probation)","Employees-on-probation grid, Start Probation, review checkpoints, days-left, decision actions",3.0,"State-dependent; needs employees seeded on probation"),
 ("Core HR","Probation Templates (/hrms/probation-templates)","New Template (duration, checkpoint days, criteria), default/active flags, edit, delete",2.0,"CRUD master feeding the probation flow"),
 ("Core HR","Probation Report (/hrms/probation-report)","Date-range + outcome filter, Run Report, Export Excel, overdue/decision columns",2.0,"Filter-first report; export not downloaded, only triggered"),
 ("Core HR","Resigned Employees (/resigned-employees)","List, name filter, columns (date/name/phone/designation/nationality)",1.0,"Read-only list"),
 ("Core HR","Employee Excel Import (/upload-employee)","Download sample, Excel Rules dialog, valid upload, malformed-file rejection, row-level errors",3.0,"File-upload path + negative fixtures; slowest to set up"),
 ("Core HR","Letter Templates (/letters/templates)","New Template, merge-fields reference, owner/type/subject, active toggle, edit, delete",2.5,"CRUD + merge-field token verification"),
 ("Core HR","Generate Letter (/letters/generate)","Template/employee pick, Preview render, Generate, send-mail option, lands in ESS letters",3.0,"Cross-module verification into /ess/letters"),
 ("Core HR","My Approvals (/approvals)","Tabs: Awaiting my decision / My requests / History; counts, approve & reject, level/raised-by columns",3.0,"3 tabs; needs pending items raised from other sub-modules"),
 ("Core HR","Approval Config (/approval/config)","New workflow (type, name), + Add level chain builder, default flag, Save Workflow, configured list",3.5,"Chain builder with variable levels; drives every approval path"),
 ("Core HR","Deduction & Remark Reports","/employee-deduction-report + /employee-remark-report: assignee & period filters, View Report",2.0,"2 filter-first report pages; empty until queried"),

 # ─────────────────── RECRUITMENT & ONBOARDING (15 pages) ───────────────────
 ("Recruitment & Onboarding","Job Requisitions (/requisition-list)","New Requisition (designation, department, positions, type, work type, branch), status, designation search",2.5,"Create form + list filters; feeds the whole hiring funnel"),
 ("Recruitment & Onboarding","Job Board / Hiring (/vacancy-list)","Add Job Opening, status filter, published/total counters, tab strip to Candidates & Talent Pools",3.0,"Tabs are router links (leave the page) — verified navigation targets"),
 ("Recruitment & Onboarding","Public careers page (/current-openings)","'Join Our Team' listing, opening selector, apply entry point, unauthenticated access check",2.0,"Public page; needs a published opening seeded to exercise apply"),
 ("Recruitment & Onboarding","Job Applications (/job-applications-list)","Applicant grid, details view, Schedule Interview action, Reject action, status transitions",2.5,"Row actions mutate candidate state; needs applications seeded"),
 ("Recruitment & Onboarding","Candidates register (/candidates)","Status buckets New/In Progress/Shortlisted/Selected/Rejected with live counts, search, branch/designation filters",3.0,"5 buckets x count integrity vs grid"),
 ("Recruitment & Onboarding","Candidate create form (/candidate/0)","Routed form: name/email/phone (+country code)/WhatsApp, branch, user type, status, skills multi-select, résumé upload",3.5,"Routed page (not a modal); file upload + multi-select"),
 ("Recruitment & Onboarding","Assessments (/assessment-list)","New Assessment (title, type, description, max score, attachment), library list, edit, delete",2.0,"CRUD + attachment"),
 ("Recruitment & Onboarding","Interview Schedules (/interview-schedules)","Schedule Interview (candidate, round, datetime, mode), status, reschedule, cancel",2.5,"Depends on interview-round master + candidates"),
 ("Recruitment & Onboarding","Offers (/offer-list)","New Offer (candidate, total CTC, joining date), status lifecycle, accept/decline",2.5,"Gateway from selection to onboarding"),
 ("Recruitment & Onboarding","Recruitment Pipeline (/recruitment-pipeline)","Kanban board, mandatory vacancy filter, Configure Stages inline editor, Score, auto-sync toggle, card drag between stages",4.5,"Kanban drag-drop + inline stage editor; heaviest UI in the module"),
 ("Recruitment & Onboarding","Communication Templates (/communication-templates)","New Template (name, type, subject, body), list, edit, delete, use in candidate comms",2.0,"CRUD master"),
 ("Recruitment & Onboarding","Talent Pool (/talent-pool)","Name/email search, skill & pool filters, tags, score, archive/restore actions",2.0,"Search-driven list; archive path from rejected candidates"),
 ("Recruitment & Onboarding","Recruitment settings","/candidate-status (status + nature) and /interview-rounds (round + order): create, reorder, edit, delete",2.0,"2 small masters that drive the funnel stages"),
 ("Recruitment & Onboarding","Onboarding Templates (/onboarding-templates)","New Template, joining checklist items, defaults, edit, delete",2.0,"CRUD master feeding onboarding runs"),
 ("Recruitment & Onboarding","Onboarding Pipeline (/onboarding-pipeline)","Start Onboarding wizard from an accepted offer, checklist progress, completion -> employee record",2.5,"Cross-module: ends by creating an employee"),

 # ────────────────────── ATTENDANCE & TIME (15 pages) ──────────────────────
 ("Attendance & Time","Shifts & Rules (/shifts)","New Shift (name, type, timing, night-shift flag), active toggle, shift-name search, edit, delete",3.0,"Master that every attendance calculation depends on"),
 ("Attendance & Time","Shift Roster (/shift-roster)","Assign shift by scope (Branch/Department/Employee/Company), date range, current assignments grid, active-only filter",3.5,"Scope cascade x4; filters live in the offcanvas panel"),
 ("Attendance & Time","Attendance Log (/attendance-log)","Filter-first report: worked/OT/balance/must-work hours, entry-exit, session details, employee status",3.0,"15 columns; grid empty until filters applied"),
 ("Attendance & Time","Data from Device (/data-from-device)","Biometric punch feed: punch type, recognition type, registered-in-system flag, device, sync time, image",2.5,"Device-fed data; needs punches seeded to assert"),
 ("Attendance & Time","Add Visit Report (/add-visit-report)","Field-visit punches: check-in/out, purpose, site name, site image, mobile location",2.5,"Geo + image columns; header misspelt 'Add Vist Report' (bug to log)"),
 ("Attendance & Time","Regularization (/regularization)","Raise correction (employee, date, type, in/out datetime, reason, attachment), request queue, approve/reject",3.5,"Create + approval queue + file attach; ESS raises into the same queue"),
 ("Attendance & Time","Overtime Approval (/overtime-approval)","OT queue: minutes, eligibility, payout, exported flag, approve/reject, search + date range",3.0,"Approval path feeding payroll; offcanvas filters"),
 ("Attendance & Time","Attendance Finalization (/attendance-finalization)","Start/refresh pay-cycle run, period/scope/target/cut-off, pending count, Finalize, run history",4.0,"State-mutating pay-cycle gate; must not finalize real periods in tests"),
 ("Attendance & Time","Geofences (/geofences)","Add Location (scope, applies-to, lat/long, radius), status, activate/deactivate, search",3.0,"Map-backed coordinates; receives approved ESS locations"),
 ("Attendance & Time","Timesheet (/timesheet)","Attendance hours vs task hours comparison, employee search, date range, shift/status filters",2.5,"Cross-module reconciliation vs Task Management"),
 ("Attendance & Time","Attendance Report Pack (/attendance-report-pack)","Daily register, report-type/branch/department/employee filters, date range, Export, Previous/Next paging",3.0,"Multi-report pack with paging; export triggered not downloaded"),
 ("Attendance & Time","Operation approval (+report)","/approval-operation and /approval-operation-report: worked/OT/balance hours approval, late/early minutes, fixed hours, remarks",3.0,"2 pages; hour-arithmetic verification"),
 ("Attendance & Time","Absent approval (+report)","/approval-absent and /approval-absent-report: absence approval by period, must-work hours, fixed hours, remarks, delete",3.0,"2 pages; mirrors the operation pair"),

 # ───────────────────── LEAVE MANAGEMENT (21 pages) ─────────────────────
 ("Leave Management","Leave Types (/leave-types)","Create type, half-day support flag, document-required flag, edit, delete, duplicate rejection",2.0,"Root master of the leave chain"),
 ("Leave Management","Leave Patterns (/leave-patterns)","New pattern, per-type entitlement details, edit, delete",2.0,"Depends on leave types"),
 ("Leave Management","Leave Policy (/leave-policy)","Choose a leave pattern, policy configuration, save, effect on assignments",2.0,"Config page linking pattern -> policy"),
 ("Leave Management","Leave Assignment (/leave-assignment-list)","New assignment (type: branch/department/employee), target, pattern, Filter, edit, delete",2.5,"Assignment matrix drives every employee's entitlement"),
 ("Leave Management","Leave Request (/leave-request-list)","New Leave Request (type, start/end, half-day, remarks), list, approval status, edit/cancel rules",2.5,"Admin-side request creation; mirrors ESS apply"),
 ("Leave Management","Leave Approval (/leave-approval)","Worklist, 3 filter selects, row select + Approve/Reject Selected (bulk), Clear, Delegate approvals modal",4.0,"Bulk decisions + delegation modal; posts to ledger & balances"),
 ("Leave Management","My Leave Policy (/my-leave-policy)","Year selector, Show, policy entitlement display for the logged-in user",1.5,"Read-only personal view"),
 ("Leave Management","Leave Balances (/leave-balances)","Year, Show, Run Accrual, columns Opening/Accrued/Carried-Fwd/Used/Encashed/Reserved/Available/Liability",3.5,"Accrual engine trigger + 10-column arithmetic"),
 ("Leave Management","Leave Ledger (/leave-ledger)","Append-only txn history, employee/type/date filters (offcanvas), balance-after integrity, Export",3.0,"Append-only guarantee + running-balance verification"),
 ("Leave Management","Leave / Attendance Sync (/leave-attendance-sync)","Year & period filters, sync status, LOP flags, attendance rows touched, Recalculate period",3.0,"Recalculation is state-mutating — assert read-only paths only"),
 ("Leave Management","Leave Encashment (/leave-encashment)","Request encashment (type, days, amount), my-requests list, status",2.0,"Request side; feeds payroll"),
 ("Leave Management","Encashment Approval (/leave-encashment-approval)","5 filter selects, request grid, approve/reject, amount verification",2.5,"Approval side of encashment"),
 ("Leave Management","Leave Delegation (/leave-delegation)","Active & past delegations registry, from/to, date range, active flag",2.0,"Created from Leave Approval; read-heavy registry"),
 ("Leave Management","Employee Handover (/employee-handover)","Set up handover (from/to employee, dates, covers), active flag, all-handovers grid, edit, delete",2.5,"Admin twin of the ESS handover page"),
 ("Leave Management","Comp-Offs (/comp-offs)","Request Comp-Off, earned/source/days/expiry, my credits grid, status",2.0,"Employee-facing credit request"),
 ("Leave Management","Comp-Off Management (/comp-off-management)","All-employee credits, pending-only toggle, Grant / Reject per row, expiry handling",2.5,"Per-row grant/reject; only present when rows exist"),
 ("Leave Management","Holidays (/holiday-list)","New Holiday (name, dates, calendar), search, Export, Calendar view at /holiday-calendar",2.5,"CRUD + separate calendar route"),
 ("Leave Management","Holiday Assignment (/holiday-assignment-list)","New assignment (type, target), Filter, list, edit, delete",2.0,"Assignment matrix for holiday calendars"),
 ("Leave Management","Leave Reports (/leave-reports)","Register / Balance / Utilization report switch, year & branch & department filters, Filter",3.0,"3 report modes behind one page"),
 ("Leave Management","Absence Analytics (/absence-analytics)","Monthly trend, absence rate by department, Bradford Factor risk table (Spells/Days/Score/Signal), filters",3.0,"Charts + Bradford scoring; collapses to 'No analytics available' with no data"),
 ("Leave Management","Leave Calendar (/leave-calendar)","Month/branch/department view, employee-name filter, Show, leave markers on calendar",2.0,"Calendar visualisation, not a grid"),

 # ─────────────── MY WORKSPACE / ESS (11 pages) ───────────────
 ("My Workspace (ESS)","ESS Dashboard (/ess)","KPI tiles (Leave Available / Today's Attendance / Pending Requests / Letters to Acknowledge), profile card, Quick Actions, Upcoming Holidays, Birthdays",2.5,"Async-loading widgets; KPI values must reconcile with source pages"),
 ("My Workspace (ESS)","My Profile (/ess/profile)","Read-only overview (code/branch/department/designation/reports-to/joined), Request A Change form (name/email/ID/passport/blood group + reason), Submit Change Request",3.0,"Write path is an approval request, not a direct edit"),
 ("My Workspace (ESS)","My Requests (/ess/requests)","Profile change request list, status tracking, documented empty state",1.5,"Round-trip verification target for /ess/profile"),
 ("My Workspace (ESS)","My Leave (/ess/leave)","Balances table (Balance/Reserved/Available), Apply For Leave (type, from/to, half-day, reason), my requests, year Show",3.0,"Primary ESS->admin E2E: apply -> approve -> ledger/balance"),
 ("My Workspace (ESS)","My Handover (/my-handover)","Set Up Handover (assignee, dates, covers, reason), Save, Clear, my-handovers grid",2.0,"Self-service twin of admin handover"),
 ("My Workspace (ESS)","My Attendance (/ess/attendance)","Date-range Show, history grid (entry/exit/worked/OT/status), Regularize, Raise OT",3.0,"Raises into admin regularization & OT queues"),
 ("My Workspace (ESS)","My Locations (/ess/locations)","Add work location (name, radius; lat/long are map-set read-only), address search, Use my location, Submit for approval, my-locations grid",3.5,"Google-Maps widget; geolocation not grantable headless; grid paints below the map"),
 ("My Workspace (ESS)","My Documents (/ess/documents)","Upload A Document (type, number, category, expiry, file), documents grid, expiry status",2.5,"File-upload path with expiry tracking"),
 ("My Workspace (ESS)","My Letters (/ess/letters)","Request a letter/certificate, published letters grid, acknowledge action, KPI decrement",2.0,"Acknowledge flips state and the dashboard KPI"),
 ("My Workspace (ESS)","My Pay (/ess/payslips)","Year Show, payslip grid (Period/Basic/Deductions/Net/Payable/Paid), PDF-unavailable footnote",2.0,"Terminal assertion of the payroll chain"),
 ("My Workspace (ESS)","My Probation (/ess/probation)","Probation status panel, review checkpoints, documented not-on-probation state",1.5,"Read-only mirror of admin probation"),

 # ────────────────────────── CROSS-CUTTING ──────────────────────────
 ("Cross-Cutting","Login & session","Company-code/username/password form, invalid-credential negatives, session persistence, logout, redirect-to-login on expiry",2.0,"Gate for every other check; 3-field tenant login"),
 ("Cross-Cutting","Navigation & menu integrity","HRMS menu tree (Core HR / Recruitment / Attendance / Leave / ESS groups), every route reachable, breadcrumb correctness",2.5,"80 routes to walk; catches broken/removed menu entries"),
 ("Cross-Cutting","Approval workflow E2E","Config a chain then drive leave, salary revision, regularization, encashment and ESS profile-change through /approvals to completion",4.0,"Multi-module chain; the highest-value regression path"),
 ("Cross-Cutting","Test data seeding & environment","Employees, shifts+roster, leave types/patterns/policy/assignment, holidays, requisition->opening->candidate, probation records",3.0,"Prerequisite for most state-dependent checks; done once per environment"),
 ("Cross-Cutting","Defect logging & retest","Log build bugs found (remark header, 'Add Vist Report' typo, upload-employee stray card title), retest after fixes",2.5,"Known issues already observed during the study"),
]

HDR = ["Project","Module","Sub-Module","Page / Functionality Check","Scope of Check","Estimation Rationale",
       "Estimated Hours","Actual Hours","Date Worked","Status","Variance (h)","Remarks"]

wb = openpyxl.Workbook()

# ---------- Sheet 1: WBS Tracker ----------
ws = wb.active; ws.title = "WBS Tracker"
hf = PatternFill("solid", fgColor="1F3864"); thin = Side(style="thin", color="C9C9C9")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
for c, h in enumerate(HDR, 1):
    cell = ws.cell(1, c, h); cell.fill = hf; cell.font = Font(bold=True, color="FFFFFF")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True); cell.border = border
for r, (sub, page, scope, est, why) in enumerate(ROWS, 2):
    vals = [P, M, sub, page, scope, why, est, None, None, "Pending", f"=G{r}-H{r}", None]
    for c, v in enumerate(vals, 1):
        cell = ws.cell(r, c, v); cell.alignment = Alignment(vertical="top", wrap_text=True); cell.border = border
    ws.cell(r, 7).number_format = "0.0"; ws.cell(r, 8).number_format = "0.0"; ws.cell(r, 11).number_format = "0.0"
n = len(ROWS) + 1
tr = n + 1
ws.cell(tr, 6, "TOTAL").font = Font(bold=True)
ws.cell(tr, 7, f"=SUM(G2:G{n})").font = Font(bold=True)
ws.cell(tr, 8, f"=SUM(H2:H{n})").font = Font(bold=True)
ws.cell(tr, 7).number_format = "0.0"; ws.cell(tr, 8).number_format = "0.0"
dv = DataValidation(type="list", formula1='"Pending,In Progress,Completed,Blocked"', allow_blank=True)
ws.add_data_validation(dv); dv.add(f"J2:J{n}")
widths = [12, 10, 24, 34, 52, 42, 11, 10, 12, 12, 11, 24]
for i, w in enumerate(widths, 1): ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = "A2"
tab = Table(displayName="WBS_HRMS", ref=f"A1:L{n}")
tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
ws.add_table(tab)

# ---------- Sheet 2: Daily Hours Log ----------
dl = wb.create_sheet("Daily Hours Log")
for c, h in enumerate(["Date","Module","Sub-Module","Hours Spent","What was done","Status Update"], 1):
    cell = dl.cell(1, c, h); cell.fill = hf; cell.font = Font(bold=True, color="FFFFFF"); cell.alignment = Alignment(horizontal="center")
for i, w in enumerate([12, 12, 24, 12, 50, 16], 1): dl.column_dimensions[get_column_letter(i)].width = w
dl.freeze_panes = "A2"
t2 = Table(displayName="DailyLogHRMS", ref="A1:F2")
t2.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
dl.add_table(t2)

# ---------- Sheet 3: Dashboard (rolled up by SUB-MODULE) ----------
db = wb.create_sheet("Dashboard")
db["A1"] = "WBS DASHBOARD — HRMS (Core HR · Recruitment · Attendance · Leave · ESS)"; db["A1"].font = Font(bold=True, size=14)
db["A3"] = "By Sub-Module"; db["A3"].font = Font(bold=True)
hdrs = ["Sub-Module","Checks","Completed","Pending/In-Prog","Est Hours","Actual Hours","Variance (h)","% Complete"]
for c, h in enumerate(hdrs, 1):
    cell = db.cell(4, c, h); cell.fill = hf; cell.font = Font(bold=True, color="FFFFFF")
SUBS = []
for r in ROWS:
    if r[0] not in SUBS: SUBS.append(r[0])
for i, s in enumerate(SUBS, 5):
    db.cell(i, 1, s)
    db.cell(i, 2, f"=COUNTIF('WBS Tracker'!C:C,A{i})")
    db.cell(i, 3, f"=COUNTIFS('WBS Tracker'!C:C,A{i},'WBS Tracker'!J:J,\"Completed\")")
    db.cell(i, 4, f"=B{i}-C{i}")
    db.cell(i, 5, f"=SUMIF('WBS Tracker'!C:C,A{i},'WBS Tracker'!G:G)")
    db.cell(i, 6, f"=SUMIF('WBS Tracker'!C:C,A{i},'WBS Tracker'!H:H)")
    db.cell(i, 7, f"=E{i}-F{i}")
    db.cell(i, 8, f"=IF(B{i}=0,0,C{i}/B{i})"); db.cell(i, 8).number_format = "0%"
TOT = 5 + len(SUBS)
db.cell(TOT, 1, "TOTAL").font = Font(bold=True)
for c in range(2, 8): db.cell(TOT, c, f"=SUM({get_column_letter(c)}5:{get_column_letter(c)}{TOT-1})").font = Font(bold=True)
db.cell(TOT, 8, f"=IF(B{TOT}=0,0,C{TOT}/B{TOT})").number_format = "0%"

r0 = TOT + 2
db.cell(r0, 1, "Daily hours (from Daily Hours Log)").font = Font(bold=True)
db.cell(r0 + 1, 1, "Today's logged hours:"); db.cell(r0 + 1, 2, "=SUMIF(DailyLogHRMS[Date],TODAY(),DailyLogHRMS[Hours Spent])")
db.cell(r0 + 2, 1, "Total logged hours:");   db.cell(r0 + 2, 2, "=SUM(DailyLogHRMS[Hours Spent])")

r1 = r0 + 5
db.cell(r1, 1, "HOW TO USE (daily routine)").font = Font(bold=True, color="9C6500")
tips = [
 "1. Each day, add a row in 'Daily Hours Log' per sub-module you worked on (Date, Module, Sub-Module, Hours).",
 "2. Update 'Actual Hours' + 'Date Worked' + 'Status' on the matching WBS Tracker row (add hours cumulatively).",
 "3. Variance auto-calculates (Estimated - Actual); negative = over-estimate risk, raise early.",
 "4. Filter the WBS Tracker Status column for Completed vs Pending (it is an Excel Table - use the header filters).",
 "5. Pivot: Insert > PivotTable > table 'WBS_HRMS' > Rows: Sub-Module, Page; Values: Sum of Estimated/Actual Hours.",
 "6. Burndown by day: Insert > PivotTable on table 'DailyLogHRMS' > Rows: Date; Values: Sum of Hours Spent.",
 "7. Estimates assume the config chain is seeded first (shifts+roster, leave types->pattern->policy->assignment,",
 "   holidays, requisition->opening->candidate) - see the 'Test data seeding' row under Cross-Cutting.",
]
for i, t in enumerate(tips, r1 + 1): db.cell(i, 1, t)
db.column_dimensions["A"].width = 34
for col in "BCDEFGH": db.column_dimensions[col].width = 14

os.makedirs(os.path.dirname(REPO), exist_ok=True)
wb.save(REPO)

subs = {}
for r in ROWS: subs[r[0]] = subs.get(r[0], 0) + r[3]
print(f"saved: {os.path.normpath(REPO)}")
print(f"rows: {len(ROWS)}")
for s, h in subs.items(): print(f"  {s:28s} {h:6.1f} h")
print(f"  {'TOTAL':28s} {sum(subs.values()):6.1f} h")
