# -*- coding: utf-8 -*-
"""WBS tracker: Project -> Module -> Sub-module -> page-wise functionality checks,
with Estimated/Actual hours, daily log, and a formula dashboard (pivot-ready)."""
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
import os

DL   = r"C:\Users\PROBOOK\Downloads\WBS_CRM_TaskManagement.xlsx"
REPO = os.path.join(os.path.dirname(__file__), "..", "docs", "excel", "WBS_CRM_TaskManagement.xlsx")

P = "Progbiz ERP"
# (Module, Sub-Module, Page/Functionality Check, Scope of check, Est h, Rationale)
ROWS = [
 # ───────────────────────────── CRM ─────────────────────────────
 ("CRM","Login","Login page functionality","Form load, valid/invalid login, Enter-key submit, password eye toggle, Forgot Password, redirect to home",2.0,"8 small checks; static form, no data setup"),
 ("CRM","Homepage","Home dashboard checks","Greeting, count cards, Create New menu, Today's Schedule, timeline/calendar entry points",2.0,"Read-only verification; no forms"),
 ("CRM","Enquiry","Add Enquiry form","Branch/date/auto-number/source fields, customer phone search, New Customer modal, Assign-To, item add (single+multiple), save/cancel",4.0,"Largest form; AJAX modal + item picker need per-field checks"),
 ("CRM","Enquiry","Conditional field logic","Followup Status options; Lead Quality shows only for In-Followup; Description for Won/Lost; Next-followup date rules",2.0,"4 status paths x field-visibility matrix"),
 ("CRM","Enquiry","Enquiry Overview page","Details display (customer/status/value/items), actions: Followup, Create Quotation, edit/delete rules before/after followup",3.0,"State-dependent rules need seeded enquiries in 2 states"),
 ("CRM","Leads Listing","Leads list & tabs","New/In-Followup/Won/Lost tabs + counts, filters, search, row actions, drill-down",3.0,"4 tabs x filter combinations; count integrity vs data"),
 ("CRM","Follow-up","Follow-up modal","Open, auto date, status options, conditional Lead Quality/Description, business value, save/cancel",3.0,"Same conditional matrix as enquiry + history write"),
 ("CRM","Follow-up","Follow-up history rules","Appears in history, latest editable/deletable, older read-only",1.5,"Needs 2+ followups seeded per lead"),
 ("CRM","Followups Listing","Followups list & tabs","Today's/Delayed/Upcoming/Non-Followup tabs + counts, filters, row Followup action",2.5,"4 tabs; date-bucketing correctness needs dated data"),
 ("CRM","Quotation","Create quotation from enquiry","Auto-fill check (all fields except Valid-Upto), items carried, editable after autofill, save to overview",3.0,"Field-by-field autofill verification"),
 ("CRM","Quotation","Quotation calculations","Rate x qty gross, discount incl/excl tax, tax, total payable; terms & conditions save",3.0,"Arithmetic verification across item combinations"),
 ("CRM","Quotation","Quotation overview & actions","Overview display, View Enquiry/View Quotation, edit, follow-up from quotation",2.0,"Cross-navigation checks"),
 ("CRM","Lead Transfer","Lead transfer flow","Filters, lead select, executive pick, transfer, assignee change verification",2.0,"End-to-end with verification query; backend flaky = buffer"),
 ("CRM","Settings","Lead Sources master","Create, duplicate rejection, edit, delete, list/search",1.5,"Simple CRUD master"),
 ("CRM","Settings","Lead Followup Status master","Create with Nature mapping (In-Followup/Won/Lost), edit, delete",1.5,"CRUD + nature-mapping effect on lifecycle"),
 ("CRM","Item Master","Item create & validations","Mandatory name/variant, blank/whitespace rejection, duplicate rejection, multi-field create",2.0,"Form + 4 negative paths"),
 ("CRM","Item Master","Item list operations","Search, edit, delete (if enabled), cancel, category link",1.5,"List ops; tenant-dependent delete"),
 ("CRM","CRM Dashboard","Reports & summary integrity","Widgets load, filters (period/type/status/source/quality), summary counts vs leads listing parity",4.0,"Cross-checking numbers across pages is slow, manual"),
 ("CRM","Customers","Customer listing & detail","List, search modes, filters, detail view, lead-stage display",2.0,"Read-heavy; several filter modes"),
 # ─────────────────────── TASK MANAGEMENT ───────────────────────
 ("Task Management","Home Dashboard","Task cards & sections","Pending/Delayed/Completed/Unscheduled cards, Today's Schedule, Running/On-Hold sections, live timer display",2.5,"Counts vs list parity + timer observation"),
 ("Task Management","Add Task","Instant task (modal)","Create New->Task, branch/type/priority/task/description/party/participants, save, appears on home + My Tasks",2.5,"Full modal pass + 2-page verification"),
 ("Task Management","Add Task","Task for Later (scheduled)","Host field & user toggle list, deadline/end-time toggles, date+time set, save, calendar/timeline presence",2.5,"Host picker + schedule fields + 2-view verification"),
 ("Task Management","Add Task","Repeat (recurring) task","Daily/Weekly/Monthly recurrence, start/end time, from/to date, save, occurrence check",2.5,"3 recurrence types; occurrence validation next-day"),
 ("Task Management","Add Task","Validation & negatives","Missing task type, missing title, invalid schedule times, clear button",1.5,"4 negative paths, inline messages"),
 ("Task Management","My Tasks","List, tabs & search","Today/Delayed/Upcoming/Unscheduled/Completed tabs + counts, columns, search, row actions",2.5,"5 tabs x count/content parity"),
 ("Task Management","Delegated Tasks","Delegated list checks","Columns incl. Assignees, status reflection after assignee actions, date filter",2.0,"Needs actions performed by another user to verify"),
 ("Task Management","Created Tasks","Created list & actions","Columns, self vs other-user tasks routing, View/Delete rules (no delete after start)",2.0,"Routing rules need 2 task ownerships"),
 ("Task Management","Unscheduled Tasks","Unscheduled page actions","List/columns, Edit, Schedule (moves to scheduled), Start Now, Delete",2.5,"4 row actions each with state verification"),
 ("Task Management","Task Details","Notes & documents","Open details panel, add note (Enter=paragraph), send note, upload image/doc/audio/video, activity log",2.5,"4 upload types + note behaviour"),
 ("Task Management","Task Details","Edit / Reschedule / Add Lead","Kebab menu: Edit task fields persist, Reschedule date/time updates calendar+timeline, Add Lead opens prefilled enquiry",2.5,"3 flows with cross-page verification"),
 ("Task Management","Lifecycle","Start-Hold-Resume-End","Start, Hold (confirm+time window), timer pause, Resume, End, moves between sections, single-active-task rule",4.0,"Most complex flow; time-window rules + known concurrency bug area"),
 ("Task Management","Calendar","Calendar checks","Scheduled task on correct date/time, updates after reschedule, admin view of participants",2.0,"Date-navigation heavy"),
 ("Task Management","Timeline","Timeline checks","Started/finished/held statuses on history, unscheduled-started status correctness",2.0,"Known bug area (status display) - re-verify"),
 ("Task Management","Daily Activity Report","Report checks","All-user rows, durations (running vs hold vs completed), date/user filters",2.0,"Duration correctness needs timed tasks"),
 ("Task Management","Multi-user","Cross-user visibility","Host-assigned & participant tasks visible to assignee; admin calendar/timeline oversight",3.0,"Requires 2nd login + task seeding both directions"),
]

HDR = ["Project","Module","Sub-Module","Page / Functionality Check","Scope of Check","Estimation Rationale",
       "Estimated Hours","Actual Hours","Date Worked","Status","Variance (h)","Remarks"]

wb = openpyxl.Workbook()

# ---------- Sheet 1: WBS Tracker ----------
ws = wb.active; ws.title = "WBS Tracker"
hf = PatternFill("solid", fgColor="1F3864"); thin = Side(style="thin", color="C9C9C9")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
for c, h in enumerate(HDR, 1):
    cell = ws.cell(1, c, h); cell.fill = hf; cell.font = Font(bold=True, color="FFFFFF")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True); cell.border = border
for r, (mod, sub, page, scope, est, why) in enumerate(ROWS, 2):
    vals = [P, mod, sub, page, scope, why, est, None, None, "Pending", f"=G{r}-H{r}", None]
    for c, v in enumerate(vals, 1):
        cell = ws.cell(r, c, v); cell.alignment = Alignment(vertical="top", wrap_text=True); cell.border = border
    ws.cell(r, 7).number_format = "0.0"; ws.cell(r, 8).number_format = "0.0"; ws.cell(r, 11).number_format = "0.0"
n = len(ROWS) + 1
# totals row
tr = n + 1
ws.cell(tr, 6, "TOTAL").font = Font(bold=True)
ws.cell(tr, 7, f"=SUM(G2:G{n})").font = Font(bold=True)
ws.cell(tr, 8, f"=SUM(H2:H{n})").font = Font(bold=True)
ws.cell(tr, 7).number_format = "0.0"; ws.cell(tr, 8).number_format = "0.0"
# status dropdown
dv = DataValidation(type="list", formula1='"Pending,In Progress,Completed,Blocked"', allow_blank=True)
ws.add_data_validation(dv); dv.add(f"J2:J{n}")
widths = [12, 16, 18, 30, 46, 40, 11, 10, 12, 12, 11, 24]
for i, w in enumerate(widths, 1): ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = "A2"
tab = Table(displayName="WBS", ref=f"A1:L{n}")
tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
ws.add_table(tab)

# ---------- Sheet 2: Daily Hours Log ----------
dl = wb.create_sheet("Daily Hours Log")
for c, h in enumerate(["Date","Module","Sub-Module","Hours Spent","What was done","Status Update"], 1):
    cell = dl.cell(1, c, h); cell.fill = hf; cell.font = Font(bold=True, color="FFFFFF"); cell.alignment = Alignment(horizontal="center")
for i, w in enumerate([12, 18, 22, 12, 50, 16], 1): dl.column_dimensions[get_column_letter(i)].width = w
dl.freeze_panes = "A2"
t2 = Table(displayName="DailyLog", ref="A1:F2")
t2.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
dl.add_table(t2)
dl["A2"] = None  # first empty log row

# ---------- Sheet 3: Dashboard ----------
db = wb.create_sheet("Dashboard")
db["A1"] = "WBS DASHBOARD — CRM + Task Management"; db["A1"].font = Font(bold=True, size=14)
db["A3"] = "By Module"; db["A3"].font = Font(bold=True)
hdrs = ["Module","Checks","Completed","Pending/In-Prog","Est Hours","Actual Hours","Variance (h)","% Complete"]
for c, h in enumerate(hdrs, 1):
    cell = db.cell(4, c, h); cell.fill = hf; cell.font = Font(bold=True, color="FFFFFF")
for i, m in enumerate(["CRM", "Task Management"], 5):
    db.cell(i, 1, m)
    db.cell(i, 2, f'=COUNTIF(\'WBS Tracker\'!B:B,A{i})')
    db.cell(i, 3, f'=COUNTIFS(\'WBS Tracker\'!B:B,A{i},\'WBS Tracker\'!J:J,"Completed")')
    db.cell(i, 4, f'=B{i}-C{i}')
    db.cell(i, 5, f'=SUMIF(\'WBS Tracker\'!B:B,A{i},\'WBS Tracker\'!G:G)')
    db.cell(i, 6, f'=SUMIF(\'WBS Tracker\'!B:B,A{i},\'WBS Tracker\'!H:H)')
    db.cell(i, 7, f'=E{i}-F{i}')
    db.cell(i, 8, f'=IF(B{i}=0,0,C{i}/B{i})'); db.cell(i, 8).number_format = "0%"
db.cell(7, 1, "TOTAL").font = Font(bold=True)
for c in range(2, 8): db.cell(7, c, f"=SUM({get_column_letter(c)}5:{get_column_letter(c)}6)").font = Font(bold=True)
db.cell(7, 8, "=IF(B7=0,0,C7/B7)").number_format = "0%"

db["A10"] = "Daily hours (from Daily Hours Log)"; db["A10"].font = Font(bold=True)
db["A11"] = "Today's logged hours:"; db["B11"] = '=SUMIF(DailyLog[Date],TODAY(),DailyLog[Hours Spent])'
db["A12"] = "Total logged hours:";   db["B12"] = '=SUM(DailyLog[Hours Spent])'

db["A15"] = "HOW TO USE (daily routine)"; db["A15"].font = Font(bold=True, color="9C6500")
tips = [
 "1. Each day, add a row in 'Daily Hours Log' per sub-module you worked on (Date, Module, Sub-Module, Hours).",
 "2. Update 'Actual Hours' + 'Date Worked' + 'Status' on the matching WBS Tracker row (add hours cumulatively).",
 "3. Variance auto-calculates (Estimated - Actual); negative = over-estimate risk, raise early.",
 "4. Filter WBS Tracker Status column for Completed vs Pending tracking (it is an Excel Table - use the header filters).",
 "5. Pivot dashboard: Insert > PivotTable > select table 'WBS' > Rows: Module, Sub-Module; Values: Sum of Estimated Hours, Sum of Actual Hours; Filter: Status.",
 "6. For a burndown by day: Insert > PivotTable on table 'DailyLog' > Rows: Date; Values: Sum of Hours Spent.",
 "7. For every NEW incoming module: copy this file, replace rows with the new module's sub-module breakdown BEFORE starting work, and commit the estimate.",
]
for i, t in enumerate(tips, 16): db.cell(i, 1, t)
db.column_dimensions["A"].width = 30
for col in "BCDEFGH": db.column_dimensions[col].width = 14

wb.save(DL); wb.save(REPO)
crm = sum(r[4] for r in ROWS if r[0] == "CRM"); tm = sum(r[4] for r in ROWS if r[0] == "Task Management")
print(f"rows={len(ROWS)}  CRM={crm}h  TaskMgmt={tm}h  TOTAL={crm+tm}h")
print("saved:", DL); print("saved:", REPO)
